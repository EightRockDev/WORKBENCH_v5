"""LP/GP capital subledger — per-deal investor tracking.

One ``lp_ledger.json`` per property folder. Tracks every named investor's
commitment, called capital, distributions received, unreturned balance,
and accrued (unpaid) preferred return. Year-by-year basis calc rolls up
to K-1-ready inputs.

Eight Rock convention (per `feedback_underwriting_conventions.md`):
  - LPs fund 100% of equity. No GP co-invest by default.
  - 8% preferred return, cumulative, non-compounded, on unreturned LP capital.
  - Three-tier waterfall: pref → return of capital → 70 LP / 30 GP residual.

Schema (`lp_ledger.json`)::

    {
        "deal_id": "Miars Farm",
        "raise_target": 5_500_000,
        "investors": [
            {
                "investor_id": "lp-001",
                "name": "Alice Smith",
                "kind": "LP",
                "commitment": 500000,
                "called_capital": 500000,
                "distributions_received": 32500,
                "notes": "subscribed 2024-03-15"
            },
            {
                "investor_id": "gp-001",
                "name": "Brian McCune",
                "kind": "GP",
                "commitment": 0,
                "called_capital": 0,
                "distributions_received": 0,
                "notes": "GP — no co-invest"
            },
            ...
        ],
        "events": [
            {
                "event_id": "evt-001",
                "type": "capital_call",
                "amount": 500000,
                "investor_id": "lp-001",
                "date": "2024-03-15",
                "notes": "initial subscription"
            },
            {
                "event_id": "evt-002",
                "type": "distribution",
                "amount": 32500,
                "investor_id": "lp-001",
                "tier": "pref",
                "date": "2024-12-31",
                "notes": "Q4 2024 pref distribution"
            },
            ...
        ],
        "as_of": "2026-05-26"
    }

Events are the source of truth. ``investor.called_capital`` and
``investor.distributions_received`` are denormalized rollups recomputed
from the event log on every save (so a bad rollup is self-healing).
"""

from __future__ import annotations

import datetime as dt
import json
import uuid
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Literal


InvestorKind = Literal["LP", "GP"]
EventType = Literal["capital_call", "distribution", "commitment_adjustment", "transfer", "writedown"]
DistributionTier = Literal["pref", "roc", "residual", "promote"]


# ---------------------------------------------------------------------------
# Data classes
# ---------------------------------------------------------------------------

@dataclass
class Investor:
    investor_id: str
    name: str
    kind: InvestorKind = "LP"
    commitment: float = 0.0
    called_capital: float = 0.0          # rollup from events
    distributions_received: float = 0.0  # rollup from events
    notes: str = ""
    # Brian 5/29 v2.0.37 — email for IR update sends.
    email: str = ""

    def to_dict(self) -> dict:
        return {
            "investor_id": self.investor_id,
            "name": self.name,
            "kind": self.kind,
            "commitment": self.commitment,
            "called_capital": self.called_capital,
            "distributions_received": self.distributions_received,
            "notes": self.notes,
            "email": self.email,
        }

    @classmethod
    def from_dict(cls, d: dict) -> "Investor":
        return cls(
            investor_id=d["investor_id"],
            name=d["name"],
            kind=d.get("kind", "LP"),
            commitment=float(d.get("commitment", 0.0)),
            called_capital=float(d.get("called_capital", 0.0)),
            distributions_received=float(d.get("distributions_received", 0.0)),
            notes=d.get("notes", ""),
            email=d.get("email", ""),
        )

    @property
    def unreturned_capital(self) -> float:
        """Capital still 'out' — called minus return-of-capital distributions.

        Approximated as called_capital minus distributions (assumes all
        distributions count against capital first; pref-only accounting is
        handled by the distribution_engine module which classifies the tier)."""
        return max(0.0, self.called_capital - self.distributions_received)


@dataclass
class LedgerEvent:
    event_id: str
    type: EventType
    amount: float
    investor_id: str
    date: str  # ISO YYYY-MM-DD
    tier: DistributionTier | None = None   # only for type=distribution
    notes: str = ""

    def to_dict(self) -> dict:
        out = {
            "event_id": self.event_id,
            "type": self.type,
            "amount": self.amount,
            "investor_id": self.investor_id,
            "date": self.date,
            "notes": self.notes,
        }
        if self.tier is not None:
            out["tier"] = self.tier
        return out

    @classmethod
    def from_dict(cls, d: dict) -> "LedgerEvent":
        return cls(
            event_id=d["event_id"],
            type=d["type"],
            amount=float(d["amount"]),
            investor_id=d["investor_id"],
            date=d["date"],
            tier=d.get("tier"),
            notes=d.get("notes", ""),
        )


@dataclass
class Ledger:
    deal_id: str
    raise_target: float = 0.0
    investors: list[Investor] = field(default_factory=list)
    events: list[LedgerEvent] = field(default_factory=list)
    as_of: str = ""

    def to_dict(self) -> dict:
        return {
            "deal_id": self.deal_id,
            "raise_target": self.raise_target,
            "investors": [i.to_dict() for i in self.investors],
            "events": [e.to_dict() for e in self.events],
            "as_of": self.as_of,
        }

    @classmethod
    def from_dict(cls, d: dict) -> "Ledger":
        return cls(
            deal_id=d.get("deal_id", ""),
            raise_target=float(d.get("raise_target", 0.0)),
            investors=[Investor.from_dict(i) for i in d.get("investors", [])],
            events=[LedgerEvent.from_dict(e) for e in d.get("events", [])],
            as_of=d.get("as_of", ""),
        )

    # ---- queries ----
    def investor(self, investor_id: str) -> Investor | None:
        return next((i for i in self.investors if i.investor_id == investor_id), None)

    @property
    def total_committed(self) -> float:
        return sum(i.commitment for i in self.investors)

    @property
    def total_called(self) -> float:
        return sum(i.called_capital for i in self.investors)

    @property
    def total_distributions(self) -> float:
        return sum(i.distributions_received for i in self.investors)

    @property
    def total_unreturned(self) -> float:
        return sum(i.unreturned_capital for i in self.investors)

    @property
    def remaining_to_raise(self) -> float:
        return max(0.0, self.raise_target - self.total_committed)

    def lps(self) -> list[Investor]:
        return [i for i in self.investors if i.kind == "LP"]

    def gps(self) -> list[Investor]:
        return [i for i in self.investors if i.kind == "GP"]


# ---------------------------------------------------------------------------
# Rollup recompute — always run after any mutation
# ---------------------------------------------------------------------------

def recompute_rollups(ledger: Ledger) -> Ledger:
    """Rebuild ``called_capital`` and ``distributions_received`` for every
    investor from the event log. Self-healing if rollups drift."""
    by_id: dict[str, Investor] = {i.investor_id: i for i in ledger.investors}
    # Zero out rollups
    for inv in by_id.values():
        inv.called_capital = 0.0
        inv.distributions_received = 0.0
    for ev in ledger.events:
        inv = by_id.get(ev.investor_id)
        if inv is None:
            continue
        if ev.type == "capital_call":
            inv.called_capital += ev.amount
        elif ev.type == "distribution":
            inv.distributions_received += ev.amount
        elif ev.type == "writedown":
            inv.called_capital = max(0.0, inv.called_capital - ev.amount)
        elif ev.type == "commitment_adjustment":
            inv.commitment += ev.amount
    ledger.as_of = dt.date.today().isoformat()
    return ledger


# ---------------------------------------------------------------------------
# Public operations
# ---------------------------------------------------------------------------

def new_id(prefix: str) -> str:
    return f"{prefix}-{uuid.uuid4().hex[:8]}"


def add_investor(
    ledger: Ledger,
    name: str,
    commitment: float,
    kind: InvestorKind = "LP",
    notes: str = "",
    email: str = "",
) -> Investor:
    inv = Investor(
        investor_id=new_id("lp" if kind == "LP" else "gp"),
        name=name,
        kind=kind,
        commitment=commitment,
        notes=notes,
        email=email,
    )
    ledger.investors.append(inv)
    return inv


def update_investor(
    ledger: Ledger,
    investor_id: str,
    *,
    name: str | None = None,
    kind: InvestorKind | None = None,
    commitment: float | None = None,
    notes: str | None = None,
    email: str | None = None,
) -> Investor | None:
    """Brian 5/29 v2.0.37 — update an existing investor's metadata.
    Returns the updated investor, or None if the id isn't found.
    Capital + distribution fields are NOT editable here (they're
    rollups from recorded events)."""
    for inv in ledger.investors:
        if inv.investor_id == investor_id:
            if name is not None: inv.name = name
            if kind is not None: inv.kind = kind
            if commitment is not None: inv.commitment = float(commitment)
            if notes is not None: inv.notes = notes
            if email is not None: inv.email = email
            return inv
    return None


def remove_investor(ledger: Ledger, investor_id: str) -> bool:
    """Brian 5/29 v2.0.37 — drop an investor from the ledger AND any
    of their events. Returns True if found+removed. Safe to call with
    an unknown id."""
    before = len(ledger.investors)
    ledger.investors = [
        inv for inv in ledger.investors
        if inv.investor_id != investor_id
    ]
    if len(ledger.investors) == before:
        return False
    # Also drop events that reference this investor — recompute rollups.
    ledger.events = [
        ev for ev in ledger.events
        if getattr(ev, "investor_id", None) != investor_id
    ]
    try:
        recompute_rollups(ledger)
    except NameError:
        pass  # Helper defined further down; safe if not yet imported
    return True


def record_capital_call(
    ledger: Ledger,
    investor_id: str,
    amount: float,
    date: str | None = None,
    notes: str = "",
) -> LedgerEvent:
    ev = LedgerEvent(
        event_id=new_id("evt"),
        type="capital_call",
        amount=amount,
        investor_id=investor_id,
        date=date or dt.date.today().isoformat(),
        notes=notes,
    )
    ledger.events.append(ev)
    recompute_rollups(ledger)
    return ev


def record_distribution(
    ledger: Ledger,
    investor_id: str,
    amount: float,
    tier: DistributionTier = "pref",
    date: str | None = None,
    notes: str = "",
) -> LedgerEvent:
    ev = LedgerEvent(
        event_id=new_id("evt"),
        type="distribution",
        amount=amount,
        investor_id=investor_id,
        tier=tier,
        date=date or dt.date.today().isoformat(),
        notes=notes,
    )
    ledger.events.append(ev)
    recompute_rollups(ledger)
    return ev


def delete_event(ledger: Ledger, event_id: str) -> bool:
    before = len(ledger.events)
    ledger.events = [e for e in ledger.events if e.event_id != event_id]
    if len(ledger.events) < before:
        recompute_rollups(ledger)
        return True
    return False


# ---------------------------------------------------------------------------
# Pref accrual math
# ---------------------------------------------------------------------------

def compute_accrued_pref(
    ledger: Ledger,
    investor_id: str,
    as_of_date: dt.date | None = None,
    pref_rate: float = 0.08,
) -> float:
    """Cumulative, non-compounded pref accrued for the investor.

    Walks the event log chronologically. Pref accrues at ``pref_rate`` on
    unreturned LP capital between events. Distributions to this investor
    reduce accrued pref before applying to capital balance — caller (i.e.
    `distribution_engine`) decides how to split distributions across pref/ROC
    tiers; this function just tells you what pref balance is OUTSTANDING.

    Returns the unpaid accrued pref as of ``as_of_date`` (default = today).
    """
    inv = ledger.investor(investor_id)
    if inv is None:
        return 0.0
    if inv.kind != "LP":
        return 0.0  # only LPs receive pref under Eight Rock structure

    target = as_of_date or dt.date.today()

    # Build a chronological event stream for this investor
    events = sorted(
        [e for e in ledger.events if e.investor_id == investor_id],
        key=lambda e: e.date,
    )
    if not events:
        return 0.0

    accrued = 0.0
    unreturned = 0.0
    prev_date = dt.date.fromisoformat(events[0].date)

    for ev in events:
        try:
            ev_date = dt.date.fromisoformat(ev.date)
        except ValueError:
            continue
        if ev_date > target:
            break
        # Accrue pref between prev_date and ev_date
        days = (ev_date - prev_date).days
        if days > 0 and unreturned > 0:
            accrued += unreturned * pref_rate * days / 365.25

        if ev.type == "capital_call":
            unreturned += ev.amount
        elif ev.type == "distribution":
            if ev.tier == "pref":
                accrued = max(0.0, accrued - ev.amount)
            else:  # roc, residual, promote all reduce unreturned capital
                unreturned = max(0.0, unreturned - ev.amount)

        prev_date = ev_date

    # Accrue from last event to as_of
    days = (target - prev_date).days
    if days > 0 and unreturned > 0:
        accrued += unreturned * pref_rate * days / 365.25

    return accrued


# ---------------------------------------------------------------------------
# IO — JSON storage in property folder
# ---------------------------------------------------------------------------

def ledger_path(folder: Path) -> Path:
    return folder / "lp_ledger.json"


def load(folder: Path) -> Ledger:
    """Load a property's ledger. Returns an empty ledger if file is missing."""
    from core.storage import get_storage
    from data.property_io import _rel
    storage = get_storage()
    key = f"{_rel(folder)}/lp_ledger.json"
    if not storage.is_file(key):
        return Ledger(deal_id=folder.name)
    try:
        data = json.loads(storage.read_text(key))
        ledger = Ledger.from_dict(data)
        # Always recompute rollups on load — self-heals
        return recompute_rollups(ledger)
    except (json.JSONDecodeError, KeyError, OSError):
        return Ledger(deal_id=folder.name)


def save(folder: Path, ledger: Ledger) -> None:
    """Persist a ledger. Always recomputes rollups before writing."""
    from core.storage import get_storage
    from data.property_io import _rel
    storage = get_storage()
    recompute_rollups(ledger)
    storage.write_text(
        f"{_rel(folder)}/lp_ledger.json",
        json.dumps(ledger.to_dict(), indent=2, default=str),
    )


# ---------------------------------------------------------------------------
# K-1 ready XLSX export
# ---------------------------------------------------------------------------

def export_k1_xlsx(ledger: Ledger, output_path: Path) -> None:
    """Write a K-1-input-ready XLSX with one sheet per investor.

    Each sheet shows:
      - Investor identity (name, kind, commitment)
      - Year-by-year capital contributions
      - Year-by-year distributions (by tier)
      - Accrued pref
      - Outside basis (commitment - distributions)
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    # Drop the default sheet — we'll add per-investor sheets below
    default = wb.active
    if default is not None:
        wb.remove(default)

    summary = wb.create_sheet("Summary")
    summary.append([
        "Investor", "Kind", "Commitment", "Called Capital",
        "Distributions Received", "Unreturned Capital", "Accrued Pref",
    ])
    for c in summary[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="0F1117")
        c.alignment = Alignment(horizontal="center")

    for inv in ledger.investors:
        accrued = compute_accrued_pref(ledger, inv.investor_id) if inv.kind == "LP" else 0.0
        summary.append([
            inv.name, inv.kind, inv.commitment, inv.called_capital,
            inv.distributions_received, inv.unreturned_capital, accrued,
        ])

        # Per-investor sheet with event detail
        sheet_name = inv.name[:31] or inv.investor_id[:31]
        # openpyxl forbids duplicate sheet names — append a suffix if needed
        suffix = 1
        original = sheet_name
        while sheet_name in [s.title for s in wb.worksheets]:
            sheet_name = f"{original[:28]}-{suffix}"
            suffix += 1
        s = wb.create_sheet(sheet_name)
        s.append(["Date", "Event", "Amount", "Tier", "Notes"])
        for c in s[1]:
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="0F1117")

        for ev in sorted([e for e in ledger.events if e.investor_id == inv.investor_id],
                          key=lambda e: e.date):
            s.append([
                ev.date,
                ev.type.replace("_", " ").title(),
                ev.amount,
                (ev.tier or "").upper(),
                ev.notes,
            ])

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
