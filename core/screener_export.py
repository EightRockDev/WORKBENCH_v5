"""Download the Property Screener's result set — CSV or Excel.

Owner ask (2026-08-27): "add a feature to this screen to download the
results into CSV/Excel."

The export is a VIEW of the rows already on screen, never a second query.
``run_screener`` produced the list, the table painted it, and both files
are built from that same list — so a download can never disagree with
what the owner just read (same rule the deal workbook follows, see
core/excel_export.py).

Two deliberate differences from the on-screen table:

  * **Raw values, not display text.** The screen prints "—" for a missing
    number and "$21,500,000" for a price. A spreadsheet column holding
    "—" stops being numeric, so a blank cell is written instead and money
    goes in as a number with a currency format on the cell.
  * **More columns.** Address, market, occupancy and the property id do
    not fit the screen but cost nothing in a file, and the id is what
    makes an exported row traceable back to the app.
"""

from __future__ import annotations

import csv
import datetime as _dt
import io
from typing import Any, Callable, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import config

# (header, key, extractor, xlsx number format). One list drives the CSV
# writer, the worksheet and the tests — a column added here shows up in
# both files, which is the only way they stay in step.
_MONEY = '"$"#,##0'
_INT = "#,##0"
_PCT1 = "0.0"


def _pct(row: dict[str, Any]) -> float | None:
    """occupancy_pct is stored as a fraction (db/schema.sql); ship percent."""
    v = row.get("occupancy_pct")
    return None if v is None else round(float(v) * 100.0, 1)


_COLUMNS: tuple[tuple[str, Callable[[dict[str, Any]], Any], str | None], ...] = (
    ("Source", lambda r: r.get("source"), None),
    ("Name", lambda r: r.get("name"), None),
    ("Address", lambda r: r.get("address"), None),
    ("City", lambda r: r.get("city"), None),
    ("State", lambda r: r.get("state"), None),
    # Zip stays text: 07030 is a New Jersey zip, not the number 7030.
    ("Zip", lambda r: None if r.get("zip") in (None, "") else str(r["zip"]), "@"),
    ("Units", lambda r: r.get("units"), _INT),
    ("Class", lambda r: r.get("asset_class"), None),
    ("Year Built", lambda r: r.get("year_built"), "0"),
    ("Occupancy %", _pct, _PCT1),
    ("Market", lambda r: r.get("market"), None),
    ("Last Sale Price", lambda r: r.get("sale_price"), _MONEY),
    ("Last Sale Date", lambda r: r.get("sale_when"), None),
    ("Owner", lambda r: r.get("owner"), None),
    ("Mgmt Co", lambda r: r.get("management_company"), None),
    ("Property ID", lambda r: r.get("property_id"), None),
)

HEADERS: tuple[str, ...] = tuple(h for h, _f, _n in _COLUMNS)

_WIDTHS = {"Source": 13, "Name": 30, "Address": 28, "City": 16, "State": 7,
           "Zip": 8, "Units": 8, "Class": 7, "Year Built": 11,
           "Occupancy %": 13, "Market": 18, "Last Sale Price": 16,
           "Last Sale Date": 14, "Owner": 30, "Mgmt Co": 26,
           "Property ID": 16}

_ACCENT = config.COLORS.get("ac", "#C8900A").lstrip("#")
_HEADER_FILL = PatternFill("solid", fgColor=f"FF{_ACCENT}")
_HEADER_FONT = Font(bold=True, color="FFFFFFFF")


def to_records(rows: Iterable[dict[str, Any]]) -> list[dict[str, Any]]:
    """The export table as plain dicts — header -> value, blanks as None."""
    out = []
    for r in rows:
        rec = {}
        for header, get, _fmt in _COLUMNS:
            v = get(r)
            rec[header] = None if v == "" else v
        out.append(rec)
    return out


def build_csv_bytes(rows: Iterable[dict[str, Any]]) -> bytes:
    """UTF-8 CSV with a BOM — without it Excel mis-reads accented names."""
    buf = io.StringIO(newline="")
    w = csv.writer(buf, lineterminator="\r\n")
    w.writerow(HEADERS)
    for rec in to_records(rows):
        w.writerow(["" if rec[h] is None else rec[h] for h in HEADERS])
    return buf.getvalue().encode("utf-8-sig")


def _active_filters(filters: dict[str, Any] | None) -> list[tuple[str, str]]:
    """Only the filters actually narrowing the search, as label/value text."""
    if not filters:
        return []
    labels = {
        "name": "Property name", "city": "City", "state": "State",
        "zip": "Zip", "owner": "Owner",
        "management_company": "Management company",
        "market": "Market / submarket", "asset_class": "Class",
        "units_min": "Units — min", "units_max": "Units — max",
        "year_min": "Year built — from", "year_max": "Year built — to",
        "price_min": "Last sale price — min", "price_max": "Last sale price — max",
        "date_from": "Sold after", "date_to": "Sold before",
        "occ_min": "Occupancy % — min", "occ_max": "Occupancy % — max",
    }
    out = []
    for key, label in labels.items():
        v = filters.get(key)
        if v is None or v == "" or v == []:
            continue
        if isinstance(v, (list, tuple)):
            v = ", ".join(str(x) for x in v)
        out.append((label, str(v)))
    return out


def build_xlsx_bytes(rows: Iterable[dict[str, Any]],
                     *, filters: dict[str, Any] | None = None,
                     generated_at: _dt.datetime | None = None) -> bytes:
    """The same table as a workbook: Results, plus a Filters provenance sheet."""
    records = to_records(rows)
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    for col, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col)].width = _WIDTHS[header]

    for i, rec in enumerate(records, start=2):
        for col, (header, _get, num_fmt) in enumerate(_COLUMNS, start=1):
            value = rec[header]
            cell = ws.cell(row=i, column=col, value=value)
            if num_fmt and value is not None:
                cell.number_format = num_fmt

    ws.freeze_panes = "A2"
    if records:
        ws.auto_filter.ref = (
            f"A1:{get_column_letter(len(HEADERS))}{len(records) + 1}")

    # Provenance: a spreadsheet that leaves the app has to say what
    # search produced it, or nobody can tell two exports apart.
    meta = wb.create_sheet("Filters")
    meta["A1"] = "Property Screener export"
    meta["A1"].font = Font(bold=True, size=13)
    stamp = generated_at or _dt.datetime.now()
    meta["A2"], meta["B2"] = "Generated", stamp.strftime("%Y-%m-%d %H:%M")
    meta["A3"], meta["B3"] = "Rows", len(records)
    meta["A4"], meta["B4"] = "Workbench version", config.WORKBENCH_VERSION
    active = _active_filters(filters)
    meta["A6"] = "Filters used" if active else "Filters used: none (all records)"
    meta["A6"].font = Font(bold=True)
    for i, (label, value) in enumerate(active, start=7):
        meta[f"A{i}"] = label
        meta[f"B{i}"] = value
    meta.column_dimensions["A"].width = 26
    meta.column_dimensions["B"].width = 40

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_filename(ext: str, *, when: _dt.datetime | None = None) -> str:
    stamp = (when or _dt.datetime.now()).strftime("%Y%m%d-%H%M")
    return f"property-screener-{stamp}.{ext}"
