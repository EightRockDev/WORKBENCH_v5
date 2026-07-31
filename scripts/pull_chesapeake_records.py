"""Pull tax + sales records for all Chesapeake Class C 20-400-unit properties
in Eight Rock's property universe. Outputs:

  1. ``chesapeake-watchlist-MMDDYYYY.csv`` and ``.xlsx`` at workbench root —
     one row per property with property identity + matched assessor record.
  2. ``chesapeake-watchlist-MMDDYYYY.md`` — Brian-readable markdown report
     with per-property cards + flagged signals (recent sales, institutional
     owners, assessment-to-sale gap).
  3. ``sales.json`` written into each property folder under Properties/
     (today only Miars-Farm-116-Chesapeake has a folder — others are written
     conditionally only when a folder exists).

Address matching strategy:
    The property record and assessor use different address conventions (record: full street
    suffix "Circle"; assessor: abbreviated "CIR"). Match on shared street-
    name tokens (excluding numbers + suffix variants) plus the largest-
    assessed parcel as the "main" record for multi-parcel complexes.

Run::

    cd python_workbench
    python scripts/pull_chesapeake_records.py
"""

from __future__ import annotations

import datetime as dt
import json
import re
import sqlite3
import sys
from pathlib import Path

# Locate workbench root from script location
_HERE = Path(__file__).resolve()
WB_ROOT = _HERE.parent.parent.parent
WORKBENCH_DB = WB_ROOT / "python_workbench" / "data" / "workbench.db"
ETL_DB = WB_ROOT / "hampton-roads-etl" / "hampton_roads.db"
PROPERTIES_DIR = WB_ROOT / "Properties"

TODAY = dt.date.today()


# ---------------------------------------------------------------------------
# Address normalization
# ---------------------------------------------------------------------------

STREET_SUFFIX_MAP = {
    "STREET": "ST", "AVENUE": "AVE", "ROAD": "RD", "DRIVE": "DR",
    "CIRCLE": "CIR", "COURT": "CT", "LANE": "LN", "PLACE": "PL",
    "TRAIL": "TR", "BOULEVARD": "BLVD", "PARKWAY": "PKWY",
    "TERRACE": "TER", "PLAZA": "PLZ", "SQUARE": "SQ", "LOOP": "LP",
    "WAY": "WAY", "PATH": "PATH",
}

PUNCT_RE = re.compile(r"[#\.,]")
SPACE_RE = re.compile(r"\s+")


# All suffix codes (normalized + raw) that should NOT count as "shared name"
_STREET_SUFFIX_TOKENS = (
    set(STREET_SUFFIX_MAP.keys())
    | set(STREET_SUFFIX_MAP.values())
    | {"N", "S", "E", "W", "NE", "NW", "SE", "SW"}
)


def _normalize_address(addr: str | None) -> tuple[str, list[str]]:
    """Return (normalized_full, street_name_tokens).

    ``street_name_tokens`` is the address with: leading number stripped,
    common street suffixes ("DR", "CIR", "ST", etc.) **removed** (not just
    abbreviated), and directional indicators removed. The remaining tokens
    are the meaningful "street name" — what distinguishes "Helensburgh"
    from "Conservancy" even though both end in "DR".

    Returns empty list if no meaningful tokens remain — caller treats this
    as "can't match by street name."
    """
    if not addr:
        return "", []
    s = PUNCT_RE.sub(" ", addr.strip().upper())
    s = SPACE_RE.sub(" ", s).strip()
    tokens = s.split()
    # Drop leading number
    if tokens and tokens[0].isdigit():
        tokens = tokens[1:]
    # Drop ALL suffix-like tokens (suffixes + directionals) — they don't
    # carry identity. We keep them out of `tokens` entirely so they can't
    # contribute to overlap matching.
    tokens = [t for t in tokens if t not in _STREET_SUFFIX_TOKENS]
    return " ".join(tokens), tokens


# ---------------------------------------------------------------------------
# Data fetchers
# ---------------------------------------------------------------------------

def fetch_prop_chesapeake() -> list[dict]:
    """Chesapeake properties Brian is tracking.

    Universe = (a) Class C 20-400 units (Eight Rock target band, per
    SUMMARY-FORMAT.md) PLUS (b) any property that already has a folder
    under Properties/ — covers off-class candidates Brian is actively
    underwriting (e.g. Miars Farm), even if they're not in the Class C list.
    """
    wb = sqlite3.connect(WORKBENCH_DB)
    wb.row_factory = sqlite3.Row

    # Class C 20-400 — the target universe
    class_c = wb.execute("""
        SELECT property_id, name, address, units, year_built, occupancy_pct,
               avg_rent, manager, management_company, owner, latitude, longitude,
               asset_class
        FROM properties
        WHERE city='Chesapeake' AND asset_class='C'
          AND units BETWEEN 20 AND 400
        ORDER BY units DESC
    """).fetchall()

    # Also include any Chesapeake property that has a folder under Properties/
    # (regardless of class) — those are properties Brian is actively tracking.
    existing_ids = {r["property_id"] for r in class_c}
    extras: list = []
    if PROPERTIES_DIR.is_dir():
        for folder in PROPERTIES_DIR.iterdir():
            if not folder.is_dir() or not folder.name.lower().endswith("-chesapeake"):
                continue
            # Try to find property record for this folder by name match
            kebab = folder.name.lower()
            # Folder convention: <name-kebab>-<units>-<city>
            # We just match by name prefix being in lower(p.name)
            # Best-effort
            row = wb.execute("""
                SELECT property_id, name, address, units, year_built, occupancy_pct,
                       avg_rent, manager, management_company, owner, latitude, longitude,
                       asset_class
                FROM properties
                WHERE city='Chesapeake'
                  AND lower(replace(name,' ','-')) = substr(?, 1, length(replace(name,' ','-')))
                LIMIT 1
            """, (kebab,)).fetchone()
            if row and row["property_id"] not in existing_ids:
                extras.append(row)
                existing_ids.add(row["property_id"])

    return [dict(r) for r in list(class_c) + extras]


def fetch_chesapeake_assessor() -> list[dict]:
    """All Chesapeake va_multifamily_inventory rows."""
    etl = sqlite3.connect(ETL_DB)
    etl.row_factory = sqlite3.Row
    return [dict(r) for r in etl.execute("""
        SELECT parcel_id, gpin, address, owner, year_built,
               class_description, property_use,
               assessed_value, land_value, improvement_value,
               acreage, land_square_footage,
               last_sale_date, last_sale_price, last_sale_buyer,
               property_zip, latest_fiscal_year
        FROM va_multifamily_inventory
        WHERE city='Chesapeake'
    """).fetchall()]


def fetch_assessment_history(parcel_id: str) -> list[tuple[int, float]]:
    """Return [(fiscal_year, assessed_value), ...] sorted ascending."""
    etl = sqlite3.connect(ETL_DB)
    rows = etl.execute("""
        SELECT fiscal_year, assessed_value
        FROM va_assessment_history
        WHERE city='Chesapeake' AND parcel_id=?
        ORDER BY fiscal_year
    """, (parcel_id,)).fetchall()
    return [(int(r[0]), float(r[1])) for r in rows if r[1] is not None]


# ---------------------------------------------------------------------------
# Matching property records → assessor parcel(s)
# ---------------------------------------------------------------------------

_NUM_PREFIX_RE = re.compile(r"^(\d+)")


def _street_number(addr: str | None) -> int | None:
    if not addr:
        return None
    m = _NUM_PREFIX_RE.match(addr.strip())
    return int(m.group(1)) if m else None


# Chesapeake assessor `class_description` values that indicate multifamily.
# (Raw codes like "3352" appear in the data; these are HR-multifamily mappings
# derived from spot-checks of known complexes.)
MULTIFAMILY_CLASS_CODES = ("3352", "3353", "3354", "3355", "3356",
                           "401", "402", "403", "404", "405", "406", "407")


def _is_multifamily_class(class_desc: str | None) -> bool:
    if not class_desc:
        return False
    s = class_desc.strip()
    return any(s.startswith(code) for code in MULTIFAMILY_CLASS_CODES)


def _name_tokens(name: str) -> set[str]:
    """Strip common multifamily filler words from an property record name and
    return the meaningful tokens. Used for owner-name match pass."""
    filler = {"AT", "OF", "ON", "THE", "AND", "OAKS", "PARK", "MANOR",
              "WOODS", "VILLAGE", "PLACE", "TRAIL", "TRAILS", "HOUSE",
              "APARTMENTS", "APTS", "II", "III"}
    s = PUNCT_RE.sub(" ", name.strip().upper())
    s = SPACE_RE.sub(" ", s).strip()
    return {t for t in s.split() if t not in filler and len(t) > 2}


def match_assessor_parcels(
    prop_addr: str,
    assessor: list[dict],
    prop_name: str = "",
) -> tuple[list[dict], str]:
    """Returns (matches, match_pass) where match_pass is "address" or "owner-name"."""
    """Find assessor parcels that match the property record.

    Two-pass matching strategy:

      Pass 1 (strict address):
        - Multifamily class code
        - Shared street-name tokens (≥1 single-word, ≥2 multi-word)
        - Street number within ±3000 of record address (covers complexes
          where leasing office number is far from parcel number)

      Pass 2 (owner-name fallback, only if pass 1 empty):
        - Multifamily class code
        - ANY token from the property record name appears in the assessor
          owner field (e.g., "Wellington at Western Branch" → assessor
          owner "WELLINGTON WESTERN BRANCH LLC" matches via "WELLINGTON").

    Returns parcels sorted by overlap desc, then assessed_value desc.
    """
    prop_norm, prop_tokens = _normalize_address(prop_addr)
    prop_set = set(prop_tokens)
    min_overlap = 2 if len(prop_tokens) >= 2 else 1
    prop_num = _street_number(prop_addr)

    # ---- Pass 1: address-based ----
    pass1: list[tuple[int, dict]] = []
    for ar in assessor:
        if not _is_multifamily_class(ar.get("class_description")):
            continue
        if not prop_tokens:
            continue
        ar_norm, ar_tokens = _normalize_address(ar["address"])
        overlap = len(prop_set & set(ar_tokens))
        if overlap < min_overlap:
            continue
        ar_num = _street_number(ar["address"])
        if prop_num is not None and ar_num is not None:
            if abs(ar_num - prop_num) > 3000:
                continue
        pass1.append((overlap, ar))

    if pass1:
        pass1.sort(key=lambda t: (-t[0], -(t[1].get("assessed_value") or 0)))
        return [c[1] for c in pass1], "address"

    # ---- Pass 2: owner-name based (fallback) ----
    if not prop_name:
        return [], "none"
    name_set = _name_tokens(prop_name)
    if not name_set:
        return [], "none"
    pass2: list[tuple[int, dict]] = []
    for ar in assessor:
        if not _is_multifamily_class(ar.get("class_description")):
            continue
        owner = (ar.get("owner") or "").upper()
        if not owner:
            continue
        hits = sum(1 for t in name_set if t in owner)
        if hits >= 1:
            pass2.append((hits, ar))
    pass2.sort(key=lambda t: (-t[0], -(t[1].get("assessed_value") or 0)))
    return [c[1] for c in pass2], "owner-name"


# ---------------------------------------------------------------------------
# Format helpers
# ---------------------------------------------------------------------------

def fmt_dollars(v: float | None) -> str:
    if v is None:
        return ""
    return f"${v:,.0f}"


def fmt_unix_ms_date(v: float | None) -> str:
    """ETL stores sale dates as unix milliseconds. Convert to ISO date."""
    if v is None or v == 0:
        return ""
    try:
        d = dt.date.fromtimestamp(float(v) / 1000.0)
        return d.isoformat()
    except (TypeError, ValueError, OSError):
        return ""


def years_ago(unix_ms: float | None) -> float | None:
    if unix_ms is None or unix_ms == 0:
        return None
    try:
        sold = dt.date.fromtimestamp(float(unix_ms) / 1000.0)
        return (TODAY - sold).days / 365.25
    except (TypeError, ValueError, OSError):
        return None


def find_folder(name: str) -> Path | None:
    """Locate the Properties/<folder> matching this property record name.
    Heuristic: kebab-case the name (dashes, no spaces) and look for any
    folder whose name starts with it."""
    if not PROPERTIES_DIR.is_dir():
        return None
    needle = re.sub(r"\s+", "-", name.strip())
    needle = re.sub(r"[^A-Za-z0-9-]", "", needle).lower()
    for child in PROPERTIES_DIR.iterdir():
        if not child.is_dir():
            continue
        cname = child.name.lower()
        # Match if the folder name starts with the kebabbed record name
        if cname.startswith(needle):
            return child
    return None


# ---------------------------------------------------------------------------
# Per-property record build
# ---------------------------------------------------------------------------

def build_property_record(
    rec: dict,
    matches: list[dict],
    match_pass: str = "address",
) -> dict:
    """Produce the consolidated row for the watchlist CSV/XLSX."""
    primary = matches[0] if matches else None
    total_assessed = sum((m.get("assessed_value") or 0) for m in matches)
    total_land = sum((m.get("land_value") or 0) for m in matches)
    total_improvement = sum((m.get("improvement_value") or 0) for m in matches)
    acreage = sum((m.get("acreage") or 0) for m in matches if m.get("acreage"))

    last_sale_price = primary.get("last_sale_price") if primary else None
    last_sale_date_ms = primary.get("last_sale_date") if primary else None
    last_sale_date = fmt_unix_ms_date(last_sale_date_ms)
    yrs_since_sale = years_ago(last_sale_date_ms)

    # Pull assessment history for the primary parcel
    history = []
    if primary:
        history = fetch_assessment_history(primary["parcel_id"])

    # Signals
    signals = []
    if last_sale_price and last_sale_price > 1_000_000:
        if yrs_since_sale is not None and yrs_since_sale < 2.0:
            signals.append(f"RECENT-SALE-{yrs_since_sale:.1f}y")
        if total_assessed and last_sale_price / total_assessed > 1.2:
            gap_pct = (last_sale_price / total_assessed - 1.0) * 100
            signals.append(f"SALE-{gap_pct:.0f}%-OVER-ASSESS")
    if primary and primary.get("owner"):
        owner_upper = (primary["owner"] or "").upper()
        if any(t in owner_upper for t in ("LLC", "LP", "LIMITED", "INC", "CORP", "CAPITAL", "PARTNERS", "TRUST", "FUND")):
            signals.append("INSTITUTIONAL-OWNER")
    if match_pass == "owner-name":
        signals.append("MATCH-VIA-OWNER-NAME-(VERIFY)")

    return {
        "prop_name": rec["name"],
        "prop_address": rec["address"],
        "units": rec["units"],
        "year_built": rec["year_built"],
        "occupancy": rec["occupancy_pct"],
        "avg_rent": rec["avg_rent"],
        "prop_owner": rec.get("owner"),
        "manager": rec.get("manager"),
        "n_assessor_parcels": len(matches),
        "primary_parcel_id": primary["parcel_id"] if primary else "",
        "primary_gpin": primary.get("gpin") if primary else "",
        "primary_address_assessor": primary["address"] if primary else "",
        "assessor_owner": primary.get("owner") if primary else "",
        "total_assessed_value": total_assessed,
        "land_value": total_land,
        "improvement_value": total_improvement,
        "acreage": acreage if acreage > 0 else None,
        "latest_fy": primary.get("latest_fiscal_year") if primary else None,
        "last_sale_date": last_sale_date,
        "last_sale_price": last_sale_price,
        "years_since_last_sale": yrs_since_sale,
        "sale_to_assessment_ratio": (
            last_sale_price / total_assessed
            if last_sale_price and total_assessed else None
        ),
        "assessment_history": history,    # list of (fy, value) tuples
        "signals": ",".join(signals),
        "_prop_property_id": rec["property_id"],
        "_matches": matches,
    }


# ---------------------------------------------------------------------------
# Output writers
# ---------------------------------------------------------------------------

def write_csv(records: list[dict], path: Path) -> None:
    import csv
    cols = [
        "prop_name", "prop_address", "units", "year_built", "occupancy",
        "avg_rent", "prop_owner", "manager", "n_assessor_parcels",
        "primary_parcel_id", "primary_gpin", "primary_address_assessor",
        "assessor_owner", "total_assessed_value", "land_value",
        "improvement_value", "acreage", "latest_fy",
        "last_sale_date", "last_sale_price",
        "years_since_last_sale", "sale_to_assessment_ratio", "signals",
    ]
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=cols)
        w.writeheader()
        for r in records:
            row = {k: r.get(k) for k in cols}
            w.writerow(row)


def write_xlsx(records: list[dict], path: Path) -> None:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Chesapeake Watchlist"

    cols = [
        ("Property", "prop_name", 30),
        ("Address (record)", "prop_address", 28),
        ("Units", "units", 7),
        ("Built", "year_built", 7),
        ("Occ%", "occupancy", 6),
        ("Avg Rent", "avg_rent", 10),
        ("Manager", "manager", 22),
        ("Assessor Owner", "assessor_owner", 30),
        ("Parcel ID", "primary_parcel_id", 16),
        ("Assessed Value", "total_assessed_value", 16),
        ("Land Value", "land_value", 14),
        ("Improvements", "improvement_value", 14),
        ("FY", "latest_fy", 6),
        ("Acres", "acreage", 8),
        ("Last Sale Date", "last_sale_date", 12),
        ("Last Sale Price", "last_sale_price", 16),
        ("Yrs Since Sale", "years_since_last_sale", 12),
        ("Sale / Assess", "sale_to_assessment_ratio", 12),
        ("Signals", "signals", 36),
    ]

    # Header
    header_fill = PatternFill("solid", fgColor="0F1117")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for i, (label, _, _) in enumerate(cols, start=1):
        c = ws.cell(row=1, column=i, value=label)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")

    # Data rows
    for r_i, rec in enumerate(records, start=2):
        for c_i, (_, key, _) in enumerate(cols, start=1):
            v = rec.get(key)
            if key == "occupancy" and v is not None:
                v = round(float(v) * 100, 1)
            elif key == "sale_to_assessment_ratio" and v is not None:
                v = round(float(v), 2)
            elif key == "years_since_last_sale" and v is not None:
                v = round(float(v), 1)
            elif key == "acreage" and v is not None:
                v = round(float(v), 2)
            ws.cell(row=r_i, column=c_i, value=v)

    # Column widths
    for i, (_, _, w) in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Highlight rows with signals
    yellow = PatternFill("solid", fgColor="FFF3CD")
    for r_i, rec in enumerate(records, start=2):
        if rec.get("signals"):
            for c_i in range(1, len(cols) + 1):
                ws.cell(row=r_i, column=c_i).fill = yellow

    ws.freeze_panes = "A2"

    # ---- second sheet: full assessment history per property ----
    ws2 = wb.create_sheet("Assessment History")
    ws2.append(["Property", "Parcel ID", "Fiscal Year", "Assessed Value"])
    for c in ws2["1:1"]:
        c.font = header_font
        c.fill = header_fill
    for rec in records:
        for fy, val in rec.get("assessment_history") or []:
            ws2.append([
                rec["prop_name"], rec["primary_parcel_id"], fy, val,
            ])
    for col_idx, w in enumerate([30, 18, 8, 16], start=1):
        ws2.column_dimensions[get_column_letter(col_idx)].width = w
    ws2.freeze_panes = "A2"

    wb.save(path)


def write_markdown(records: list[dict], path: Path) -> None:
    lines: list[str] = []
    lines.append("# Chesapeake Watchlist — Tax & Sales Records")
    lines.append("")
    lines.append(f"Generated {TODAY.isoformat()} from property records + Chesapeake city assessor open data.")
    lines.append("")
    lines.append(
        f"**{len(records)} Class C properties** in Eight Rock's Chesapeake universe (20-400 units). "
        "Match heuristic: shared street-name tokens between record address and assessor parcel; "
        "for multi-parcel complexes, the largest-assessed-value parcel is shown as primary."
    )
    lines.append("")

    # Summary section
    matched = [r for r in records if r["primary_parcel_id"]]
    unmatched = [r for r in records if not r["primary_parcel_id"]]
    recent_sales = [r for r in records if r["years_since_last_sale"] is not None and r["years_since_last_sale"] < 3.0]
    institutional = [r for r in records if "INSTITUTIONAL-OWNER" in (r["signals"] or "")]

    lines.append("## Summary")
    lines.append("")
    lines.append(f"- Matched to assessor records: **{len(matched)}/{len(records)}**")
    lines.append(f"- Recently sold (< 3 years): **{len(recent_sales)}**")
    lines.append(f"- Institutional ownership signal: **{len(institutional)}**")
    if unmatched:
        lines.append(f"- Unmatched (no street-name overlap found): {len(unmatched)} — investigate manually")
    lines.append("")

    # Recently-sold callout
    if recent_sales:
        lines.append("## Recently Sold (last 3 years)")
        lines.append("")
        lines.append("These are the highest-signal entries in Brian's universe — recent sale data sets the comp.")
        lines.append("")
        for r in sorted(recent_sales, key=lambda r: r["years_since_last_sale"] or 99):
            ppu = (r["last_sale_price"] or 0) / r["units"] if r["units"] else 0
            lines.append(
                f"- **{r['prop_name']}** ({r['units']}u, {r['year_built'] or '?'}) — "
                f"sold {r['last_sale_date']} for **{fmt_dollars(r['last_sale_price'])}** "
                f"(${ppu:,.0f}/unit) to {r['assessor_owner'] or '?'}. "
                f"Sale/Assess: {r['sale_to_assessment_ratio']:.2f}x"
                if r['sale_to_assessment_ratio'] else
                f"- **{r['prop_name']}** — sold {r['last_sale_date']} for **{fmt_dollars(r['last_sale_price'])}**"
            )
        lines.append("")

    # Per-property cards
    lines.append("## Per-Property Detail")
    lines.append("")
    for r in records:
        lines.append(f"### {r['prop_name']}")
        lines.append("")
        lines.append(f"- **record address:** {r['prop_address']}")
        lines.append(f"- **Units:** {r['units']} | **Built:** {r['year_built'] or '?'} | **Occ:** {(r['occupancy'] or 0)*100:.1f}% | **Avg rent:** {fmt_dollars(r['avg_rent'])}")
        lines.append(f"- **Record owner:** {r['prop_owner'] or '?'} | **Manager:** {r['manager'] or '?'}")
        if r["primary_parcel_id"]:
            lines.append(f"- **Assessor parcel:** {r['primary_parcel_id']} at {r['primary_address_assessor']}")
            lines.append(f"- **Assessor owner:** {r['assessor_owner'] or '?'}")
            lines.append(f"- **Assessed FY{r['latest_fy']}:** {fmt_dollars(r['total_assessed_value'])} (land {fmt_dollars(r['land_value'])} + improvements {fmt_dollars(r['improvement_value'])})")
            if r["n_assessor_parcels"] > 1:
                lines.append(f"  - Note: {r['n_assessor_parcels']} parcels matched; values aggregated.")
            if r["last_sale_price"]:
                lines.append(
                    f"- **Last sale:** {r['last_sale_date']} for {fmt_dollars(r['last_sale_price'])} "
                    f"({r['years_since_last_sale']:.1f}y ago) → "
                    f"{(r['sale_to_assessment_ratio'] or 0):.2f}x assessment"
                )
            else:
                lines.append("- **Last sale:** none on record")
            hist = r.get("assessment_history") or []
            if hist:
                lines.append(
                    "- **Assessment history:** "
                    + " | ".join(f"FY{fy} {fmt_dollars(v)}" for fy, v in hist)
                )
        else:
            lines.append("- **Assessor record:** no match — manual lookup needed")
        if r["signals"]:
            lines.append(f"- **Signals:** `{r['signals']}`")
        lines.append("")

    path.write_text("\n".join(lines), encoding="utf-8")


def write_sales_json_for_property_folder(record: dict, folder: Path) -> None:
    """Append/replace the sales.json in a property folder with the most-
    recent sale + a notes line covering full assessment history."""
    if not record["primary_parcel_id"]:
        return
    sales_path = folder / "sales.json"

    notes_parts = [
        f"Chesapeake parcel {record['primary_parcel_id']}",
    ]
    if record["primary_gpin"]:
        notes_parts.append(f"GPIN {record['primary_gpin']}")
    notes_parts.append(f"Address: {record['primary_address_assessor']}")
    if record["acreage"]:
        notes_parts.append(f"{record['acreage']:.2f} acres")
    notes_parts.append(
        f"Assessor owner: {record['assessor_owner'] or '?'}"
    )
    notes_parts.append(
        f"FY{record['latest_fy']} assessed: {fmt_dollars(record['total_assessed_value'])}"
    )
    history = record.get("assessment_history") or []
    if history and len(history) > 1:
        notes_parts.append(
            "History: "
            + " | ".join(f"FY{fy} {fmt_dollars(v)}" for fy, v in history)
        )
    notes = ". ".join(notes_parts)

    if record["last_sale_price"]:
        entry = {
            "date": record["last_sale_date"],
            "price": fmt_dollars(record["last_sale_price"]),
            "grantor": "",
            "grantee": record["assessor_owner"] or "",
            "notes": notes,
        }
        sales_data = [entry]
    else:
        sales_data = [{
            "date": "",
            "price": "",
            "grantor": "",
            "grantee": "",
            "notes": notes + ". No sale on record.",
        }]

    sales_path.write_text(
        json.dumps(sales_data, indent=2),
        encoding="utf-8",
    )


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> int:
    print("Fetching Chesapeake Class C 20-400 unit properties...")
    prop_rows = fetch_prop_chesapeake()
    print(f"  found {len(prop_rows)} property records")

    print("Fetching Chesapeake assessor inventory...")
    assessor = fetch_chesapeake_assessor()
    print(f"  found {len(assessor)} assessor records")
    print()

    records: list[dict] = []
    folders_updated = 0
    for src in prop_rows:
        matches, match_pass = match_assessor_parcels(src["address"], assessor, prop_name=src["name"])
        rec = build_property_record(src, matches, match_pass=match_pass)
        records.append(rec)
        match_str = f"{rec['n_assessor_parcels']} match" if rec['n_assessor_parcels'] != 1 else "1 match"
        print(
            f"  {src['name'][:36]:<36} u={src['units']:>3}  "
            f"{match_str:<10}  "
            f"assessed={fmt_dollars(rec['total_assessed_value']):>14}  "
            f"sale={fmt_dollars(rec['last_sale_price']):>14}  "
            f"{rec['signals']}"
        )

        # Write sales.json into property folder if one exists
        folder = find_folder(src["name"])
        if folder is not None:
            write_sales_json_for_property_folder(rec, folder)
            folders_updated += 1
            print(f"    -> updated {folder.name}/sales.json")

    print()
    print(f"Folders updated: {folders_updated}")

    # Outputs at workbench root
    suffix = TODAY.strftime("%m%d%Y")
    csv_path = WB_ROOT / f"chesapeake-watchlist-{suffix}.csv"
    xlsx_path = WB_ROOT / f"chesapeake-watchlist-{suffix}.xlsx"
    md_path = WB_ROOT / f"chesapeake-watchlist-{suffix}.md"

    write_csv(records, csv_path)
    write_xlsx(records, xlsx_path)
    write_markdown(records, md_path)

    print()
    print(f"Wrote:")
    print(f"  {csv_path.name}")
    print(f"  {xlsx_path.name}")
    print(f"  {md_path.name}")
    print()
    return 0


if __name__ == "__main__":
    sys.exit(main())
