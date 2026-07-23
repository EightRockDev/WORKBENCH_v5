"""ALN Property Export(s) -> SQLite.

Reads the ALN xlsx exports, maps columns to our schema with type coercion,
and writes to `data/workbench.db`. Idempotent: drops and recreates the
`properties` table on every run.

Brian 5/30: extended from a single Virginia file to the **multi-state**
ALN library in `00-Technology/ALN Data and Reports/` (VA, NC, SC, TN, GA +
Atlanta). All modern ALN exports share the same 54-column schema with an
`API Id` UUID primary key. We load every state file, dedupe by that UUID
(a property in both the VA file and the Norfolk file is one property), tag
each record's asset type, and stash the full original row as JSON in
`raw_row` for debugging.

The custom-property merge (live deals like Crossroads, from
`Properties/_custom_props.json`) is preserved on every sync so the active
pipeline is never wiped by a rebuild.
"""

from __future__ import annotations

import datetime as dt
import json
import sqlite3
from pathlib import Path
from typing import Any, Callable

import openpyxl

# ---------------------------------------------------------------------------
# Source data location.
#
# The ALN library lives one level ABOVE the workbench root, in
# `00-Technology/ALN Data and Reports/`. `_DATA_DIR` here is
# `python_workbench/data`, so the workbench root is parent.parent and the
# ALN library is parent.parent.parent / "ALN Data and Reports".
# ---------------------------------------------------------------------------

_THIS = Path(__file__).resolve()
_WORKBENCH_ROOT = _THIS.parent.parent.parent          # ...\8-ROCK-WORKBNCH
_TECH_ROOT = _WORKBENCH_ROOT.parent                   # ...\00-Technology
ALN_DATA_DIR = _TECH_ROOT / "ALN Data and Reports"

# Back-compat: the single canonical VA file (still used as a fallback +
# referenced by db.ALN_PATH). Note the **double space** in the vendor name.
ALN_FILENAME = "ALN Virginia  Property Export - March 10th, 2026.xlsx"
ALN_SHEET_NAME = "ALN Property Data"

# The set of modern multi-state exports to ingest. All carry the 54-column
# schema with an `API Id` UUID. Dedup by that UUID handles all overlaps
# (Norfolk ⊂ VA; the GA full file ⊂ Atlanta1+Atlanta2+GA-not-Atlanta; etc.),
# so listing redundant files here is harmless — it only adds coverage.
#
# Legacy 2022/2023 files (VABEACH 2022, Artcraft, Seminole, Norfolk<15) use
# the OLD ALN schema with no `API Id` and are intentionally EXCLUDED — the
# modern exports supersede them with fresher data + owner columns.
MULTISTATE_FILENAMES: tuple[str, ...] = (
    "ALN Virginia  Property Export - March 10th, 2026.xlsx",
    "North Carolina - ALN Export March 2026.xlsx",
    "South Carolina - ALN Property Export - March 2026.xlsx",
    "Tennessee - ALN Property Export - March 2026.xlsx",
    "Georgia - ALN Property Export - March 2026.xlsx",
    "ALN Property Export - Georgia Not Atlanta - March 2026.xlsx",
    "Atlanta 1 ALN Property Export - March 2026.xlsx",
    "Atlanta 2 ALN Property - March 2026.xlsx",
    "ALN Export - Norfolk - March 2026.xlsx",
    "ALN Property Export (Virginia).xlsx",
    "ALN Property Export(3).xlsx",
)


def multistate_paths() -> list[Path]:
    """Resolve the multi-state file list to existing paths (skips missing)."""
    out = []
    for name in MULTISTATE_FILENAMES:
        p = ALN_DATA_DIR / name
        if p.is_file():
            out.append(p)
    return out


# ---------------------------------------------------------------------------
# Type coercion helpers — ALN cells can be ints, floats, strings, or None.
# Each helper returns None for missing/invalid input rather than raising,
# so a single bad cell doesn't kill the whole sync.
# ---------------------------------------------------------------------------

def _to_str(v: Any) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def _to_int(v: Any) -> int | None:
    if v is None or v == "":
        return None
    try:
        return int(float(str(v).replace(",", "").strip()))
    except (ValueError, TypeError):
        return None


def _to_float(v: Any) -> float | None:
    if v is None or v == "":
        return None
    try:
        return float(str(v).replace(",", "").replace("$", "").strip())
    except (ValueError, TypeError):
        return None


def _to_occupancy_fraction(v: Any) -> float | None:
    """ALN reports occupancy as 0-100 with possible 'N/A' text. Convert to
    a 0.0-1.0 fraction so downstream math (vacancy = 1 - occ) just works."""
    if v is None or v == "":
        return None
    s = str(v).strip().upper()
    if s in ("N/A", "NA", "-", "--"):
        return None
    try:
        x = float(s.replace("%", ""))
    except (ValueError, TypeError):
        return None
    if x > 1.0:
        x = x / 100.0
    return x


# ---------------------------------------------------------------------------
# Asset-type + segment derivation (Brian 5/30).
#
# ALN is apartment data, so ~everything is multifamily. `asset_type` defaults
# to "Multifamily" and is set precisely when a non-MF signal appears (his
# instruction: "tag exactly what it is — medical, retail, commercial, office,
# storage, land lease"). `property_segment` is the genuinely useful
# multifamily sub-distinction (you underwrite affordable / senior / student
# very differently from conventional market-rate).
# ---------------------------------------------------------------------------

def _derive_asset_type(prop_type: Any, tags: Any, name: Any) -> str:
    blob = " ".join(str(x or "").lower() for x in (prop_type, tags, name))
    if any(k in blob for k in ("self storage", "self-storage", "storage facility")):
        return "Storage"
    if any(k in blob for k in ("medical office", "medical center", "office building", "office park")):
        return "Office/Medical"
    if any(k in blob for k in ("shopping center", "strip mall", "retail center", "retail")):
        return "Retail"
    if any(k in blob for k in ("hotel", "motel", "hospitality", "extended stay")):
        return "Hospitality"
    if any(k in blob for k in ("manufactured", "mobile home", "mh community", "land lease", "ground lease", "rv park")):
        return "Manufactured/Land-Lease"
    return "Multifamily"


def _derive_segment(prop_type: Any, tags: Any, name: Any, status: Any) -> str:
    blob = " ".join(str(x or "").lower() for x in (prop_type, tags, name, status))
    if any(k in blob for k in ("section 8", "lihtc", "tax credit", "affordable", "subsidized", "hud", "section 42")):
        return "Affordable/Subsidized"
    if any(k in blob for k in ("senior", "55+", "62+", "age restricted", "age-restricted", "independent living", "assisted living")):
        return "Senior"
    if any(k in blob for k in ("student", "university", "college campus")):
        return "Student"
    if any(k in blob for k in ("military", "base housing")):
        return "Military"
    return "Conventional"


# ---------------------------------------------------------------------------
# Column map: ALN header -> (schema column, coercion function)
# Verified against the actual March-2026 exports (all 54 columns).
# ---------------------------------------------------------------------------

ALN_COLUMN_MAP: dict[str, tuple[str, Callable[[Any], Any]]] = {
    "API Id":              ("property_id",        _to_str),
    "ALN Id":              ("aln_id",             _to_str),
    "Property Name":       ("name",               _to_str),
    "Address":             ("address",            _to_str),
    "City":                ("city",               _to_str),
    "State":               ("state",              _to_str),
    "ZIP":                 ("zip",                _to_str),
    "County":              ("county",             _to_str),
    "# of Units":          ("units",              _to_int),
    "# Units":             ("units",              _to_int),   # legacy header
    "Year Built":          ("year_built",         _to_int),
    "Remodeled":           ("last_remodel",       _to_int),
    "Occupancy":           ("occupancy_pct",      _to_occupancy_fraction),
    "Average Sqft":        ("avg_sqft",           _to_float),
    "Average SqFt":        ("avg_sqft",           _to_float),  # legacy header
    "Average Rate":        ("avg_rent",           _to_float),
    "Average Rent":        ("avg_rent",           _to_float),  # legacy header
    "Average Rent/Sqft":   ("rent_per_sqft",      _to_float),
    "Avg Rent/SqFt":       ("rent_per_sqft",      _to_float),  # legacy header
    "ALN Price Class":     ("asset_class",        _to_str),
    "Property Type":       ("property_type",      _to_str),
    "Prop Type":           ("property_type",      _to_str),    # legacy header
    "Market":              ("market",             _to_str),
    "Market Description":  ("market_description", _to_str),
    "Submarket":           ("submarket",          _to_str),
    "Sub Market":          ("submarket",          _to_str),    # legacy header
    "Latitude":            ("latitude",           _to_float),
    "Longitude":           ("longitude",          _to_float),
    "Owner Name":          ("owner",              _to_str),
    "Owner Address":       ("owner_address",      _to_str),
    "Owner Phone #":       ("owner_phone",        _to_str),
    "Owner Fax #":         ("owner_fax",          _to_str),
    "Current Manager":     ("manager",            _to_str),
    "Area Supervisor":     ("area_supervisor",    _to_str),
    "Management Company":  ("management_company", _to_str),
    "Corp Mgmt Id":        ("corp_mgmt_id",       _to_str),
    "PM Software":         ("pm_software",        _to_str),
    "Asset or Fee":        ("asset_or_fee",       _to_str),
    "Lease Terms":         ("lease_terms",        _to_str),
    "Property Tags":       ("tags",               _to_str),
    "Status Description":  ("status",             _to_str),
    "Phone #":             ("property_phone",     _to_str),
    "Web Site":            ("website",            _to_str),
    "URL":                 ("website",            _to_str),     # legacy header
    "EMail Address":       ("email",              _to_str),
    # Enrichment columns present only in some older Norfolk exports — mapped
    # so they populate when available; NULL otherwise (filled by city feeds).
    "Last Sold Year":      ("last_sold_year",     _to_int),
    "Last Sold Amount":    ("last_sold_amount",   _to_float),
    "Last Sold Per Unit":  ("last_sold_per_unit", _to_float),
    "Assessed Value Per Unit": ("assessed_value_per_unit", _to_float),
}

# Schema columns in canonical insertion order. Loader produces dicts keyed by
# these names; rows missing a key get NULL.
SCHEMA_COLUMNS: tuple[str, ...] = (
    "property_id", "aln_id", "name", "address", "city", "state", "zip", "county",
    "units", "year_built", "last_remodel", "occupancy_pct",
    "avg_sqft", "avg_rent", "rent_per_sqft",
    "asset_class", "property_type", "asset_type", "property_segment",
    "market", "market_description", "submarket",
    "latitude", "longitude",
    "owner", "owner_address", "owner_phone", "owner_fax",
    "manager", "area_supervisor", "management_company", "corp_mgmt_id", "pm_software",
    "asset_or_fee", "lease_terms", "tags", "status",
    "property_phone", "website", "email",
    "last_sold_year", "last_sold_amount", "last_sold_per_unit",
    "assessed_value_per_unit",
    "source_file", "aln_pull_date", "raw_row",
)


# ---------------------------------------------------------------------------
# Loader functions
# ---------------------------------------------------------------------------

def _completeness(rec: dict[str, Any]) -> int:
    """Count non-null mapped fields — used to keep the richest row on dedup."""
    return sum(1 for k, v in rec.items() if k not in ("raw_row", "source_file") and v not in (None, ""))


def _read_sheet(path: Path, sheet_name: str, pull_date: str) -> list[dict[str, Any]]:
    """Read one worksheet into a list of schema-keyed dicts."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return []
    ws = wb[sheet_name]
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        wb.close()
        return []

    header_to_index: dict[str, int] = {
        str(name).strip(): idx for idx, name in enumerate(header) if name is not None
    }
    # Skip sheets that aren't ALN property data (e.g., the Cover sheet)
    if "Property Name" not in header_to_index and "ALN Id" not in header_to_index:
        wb.close()
        return []

    out: list[dict[str, Any]] = []
    for raw_row in rows_iter:
        if all(c is None or c == "" for c in raw_row):
            continue
        original = {
            str(h).strip(): raw_row[i] if i < len(raw_row) else None
            for h, i in header_to_index.items()
        }
        rec: dict[str, Any] = {col: None for col in SCHEMA_COLUMNS}
        for aln_header, (schema_col, coerce) in ALN_COLUMN_MAP.items():
            idx = header_to_index.get(aln_header)
            if idx is None or idx >= len(raw_row):
                continue
            val = coerce(raw_row[idx])
            # Don't let a NULL legacy-header alias overwrite a populated value
            if val is not None or rec.get(schema_col) is None:
                rec[schema_col] = val

        # Drop header-echo rows: some ALN sheets repeat the header inside the
        # data block. Such a row has literal column-name values.
        if (rec.get("name") in ("Property Name", "ALN Id", "Status")
                or rec.get("property_id") in ("API Id", "ALN Id")
                or (rec.get("state") == "State")):
            continue

        # Primary key: API Id UUID. Fallback to aln-<numeric id> so rows from
        # any export that lacks API Id are still ingested (never silently
        # dropped — that was a latent bug in the single-file loader).
        if not rec.get("property_id"):
            if rec.get("aln_id"):
                rec["property_id"] = f"aln-{rec['aln_id']}"
            else:
                continue  # truly unidentifiable row
        if not rec.get("name"):
            continue

        rec["asset_type"] = _derive_asset_type(
            rec.get("property_type"), rec.get("tags"), rec.get("name")
        )
        rec["property_segment"] = _derive_segment(
            rec.get("property_type"), rec.get("tags"), rec.get("name"), rec.get("status")
        )
        rec["source_file"] = path.name
        rec["aln_pull_date"] = pull_date
        rec["raw_row"] = json.dumps(original, default=str, ensure_ascii=False)
        out.append(rec)

    wb.close()
    return out


def load_aln_xlsx(path: Path) -> list[dict[str, Any]]:
    """Back-compat single-file loader: read the canonical ALN data sheet."""
    if not path.is_file():
        raise FileNotFoundError(f"ALN file not found: {path}")
    pull_date = dt.date.today().isoformat()
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheets = [s for s in wb.sheetnames if s.strip().lower() != "cover"]
    wb.close()
    rows: list[dict[str, Any]] = []
    seen: dict[str, dict[str, Any]] = {}
    for sh in sheets:
        for rec in _read_sheet(path, sh, pull_date):
            pid = rec["property_id"]
            if pid not in seen or _completeness(rec) > _completeness(seen[pid]):
                seen[pid] = rec
    return list(seen.values())


def load_aln_multi(paths: list[Path] | None = None) -> list[dict[str, Any]]:
    """Load every multi-state ALN export, deduped by property_id (API Id UUID).

    On a duplicate UUID across files, keep the row with the most populated
    fields. Returns one dict per unique property.
    """
    if paths is None:
        paths = multistate_paths()
    pull_date = dt.date.today().isoformat()
    seen: dict[str, dict[str, Any]] = {}
    for path in paths:
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            sheets = [s for s in wb.sheetnames if s.strip().lower() != "cover"]
            wb.close()
        except Exception as e:
            print(f"[aln_loader] could not open {path.name}: {e}")
            continue
        for sh in sheets:
            for rec in _read_sheet(path, sh, pull_date):
                pid = rec["property_id"]
                if pid not in seen or _completeness(rec) > _completeness(seen[pid]):
                    seen[pid] = rec
    return list(seen.values())


def write_to_sqlite(
    rows: list[dict[str, Any]],
    db_path: Path,
    schema_sql: Path,
) -> int:
    """Drop and recreate `properties` table, then bulk-insert all rows."""
    if not schema_sql.is_file():
        raise FileNotFoundError(f"Schema SQL not found: {schema_sql}")

    db_path.parent.mkdir(parents=True, exist_ok=True)
    schema = schema_sql.read_text(encoding="utf-8")

    conn = sqlite3.connect(db_path)
    try:
        conn.execute("DROP TABLE IF EXISTS properties")
        conn.executescript(schema)
        if not rows:
            conn.commit()
            return 0
        placeholders = ", ".join("?" for _ in SCHEMA_COLUMNS)
        col_list = ", ".join(SCHEMA_COLUMNS)
        sql = f"INSERT OR REPLACE INTO properties ({col_list}) VALUES ({placeholders})"
        values = [tuple(r.get(c) for c in SCHEMA_COLUMNS) for r in rows]
        conn.executemany(sql, values)
        conn.commit()
    finally:
        conn.close()
    return len(rows)


def sync(
    aln_path: Path | None,
    db_path: Path,
    schema_sql: Path,
    properties_root: Path | None = None,
) -> int:
    """End-to-end: load all ALN state files + custom props + write.

    `aln_path` is retained for back-compat. When the multi-state library
    folder exists, we load the full library; otherwise we fall back to the
    single `aln_path` file. Custom properties (live deals) are always merged
    so a rebuild never wipes the active pipeline.
    """
    paths = multistate_paths()
    if paths:
        rows = load_aln_multi(paths)
    elif aln_path is not None and Path(aln_path).is_file():
        rows = load_aln_xlsx(Path(aln_path))
    else:
        rows = []

    # Merge in custom properties (user-added live deals; not in ALN xlsx)
    try:
        from data.property_io import PROPERTIES_ROOT as _LIVE_ROOT
        from data.property_io import load_custom_props
        root = properties_root if properties_root is not None else _LIVE_ROOT
        custom = load_custom_props(root)
        # Custom props override ALN rows on the same property_id (they're the
        # analyst's curated truth for live deals).
        by_id = {r["property_id"]: r for r in rows}
        for cp in custom:
            row = {col: cp.get(col) for col in SCHEMA_COLUMNS}
            if not row.get("property_id") or not row.get("name"):
                continue
            # Live deals are multifamily unless explicitly tagged otherwise.
            if not row.get("asset_type"):
                row["asset_type"] = _derive_asset_type(
                    row.get("property_type"), row.get("tags"), row.get("name")
                )
            if not row.get("property_segment"):
                row["property_segment"] = _derive_segment(
                    row.get("property_type"), row.get("tags"),
                    row.get("name"), row.get("status"),
                )
            row.setdefault("aln_pull_date", dt.date.today().isoformat())
            if not row.get("raw_row"):
                row["raw_row"] = json.dumps(cp, default=str)
            by_id[row["property_id"]] = row
        rows = list(by_id.values())
    except Exception as e:
        print(f"[aln_loader.sync] custom-prop merge failed: {e}")

    return write_to_sqlite(rows, db_path, schema_sql)
