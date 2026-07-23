"""Tests for data.aln_loader — ALN xlsx → SQLite ingestion.

Strategy:
  - Build a tiny in-memory xlsx fixture mirroring the real ALN structure
    (Cover sheet + ALN Property Data sheet with the canonical 54 headers).
  - Test type coercion at the unit level (N/A, blanks, weird strings).
  - Test load + write end-to-end against a temp SQLite db.
  - Test the de-dup behavior (vendor occasionally exports duplicate rows).
"""

from __future__ import annotations

import json
import sqlite3
from pathlib import Path

import openpyxl
import pytest

from data.aln_loader import (
    SCHEMA_COLUMNS,
    _to_float,
    _to_int,
    _to_occupancy_fraction,
    _to_str,
    load_aln_xlsx,
    sync,
    write_to_sqlite,
)


# ---------------------------------------------------------------------------
# Type coercion unit tests
# ---------------------------------------------------------------------------


@pytest.mark.parametrize("v,expected", [
    (None, None),
    ("", None),
    ("hello", "hello"),
    ("  spaced  ", "spaced"),
    (123, "123"),
    (1.5, "1.5"),
])
def test_to_str(v, expected):
    assert _to_str(v) == expected


@pytest.mark.parametrize("v,expected", [
    (None, None),
    ("", None),
    ("not a number", None),
    ("1974", 1974),
    ("1974.0", 1974),
    (1974, 1974),
    (1974.7, 1974),    # truncates floats
    ("1,200", 1200),   # commas tolerated
    ("  316  ", 316),
])
def test_to_int(v, expected):
    assert _to_int(v) == expected


@pytest.mark.parametrize("v,expected", [
    (None, None),
    ("", None),
    ("garbage", None),
    ("1431.66667", pytest.approx(1431.66667)),
    (1431.66667, pytest.approx(1431.66667)),
    ("1,200.50", pytest.approx(1200.50)),
])
def test_to_float(v, expected):
    result = _to_float(v)
    if expected is None:
        assert result is None
    else:
        assert result == expected


@pytest.mark.parametrize("v,expected", [
    (None, None),
    ("", None),
    ("N/A", None),
    ("n/a", None),
    ("--", None),
    ("100", 1.0),         # 100% → 1.0
    (100, 1.0),
    ("95", 0.95),
    ("95.5", 0.955),
    ("95%", 0.95),
    (0.95, 0.95),         # already a fraction → keep
    (1.0, 1.0),
])
def test_to_occupancy_fraction(v, expected):
    result = _to_occupancy_fraction(v)
    if expected is None:
        assert result is None
    else:
        assert result == pytest.approx(expected)


# ---------------------------------------------------------------------------
# Fixture builder — in-memory xlsx that mirrors the real ALN shape
# ---------------------------------------------------------------------------


def _build_fixture_xlsx(tmp_path: Path, rows: list[dict]) -> Path:
    """Create an xlsx with a Cover sheet (skipped) and ALN Property Data sheet.

    `rows` is a list of dicts keyed by ALN header names. Missing headers get
    None. The fixture preserves the canonical column order so the loader can
    find columns by header lookup.
    """
    # Canonical header list (verified against the real ALN export 2026-05-06)
    headers = [
        "ALN Id", "Status", "Status Description", "Property Name", "Address",
        "City", "State", "ZIP", "County", "Phone #", "Fax #", "Market",
        "Market Description", "Submarket", "Lease Terms", "# of Units",
        "Year Built", "Remodeled", "Occupancy", "Average Sqft", "Average Rate",
        "Average Rent/Sqft", "ALN Price Class", "EMail Address", "Web Site",
        "Current Manager", "Manager First Name", "Manager Last Name",
        "Area Supervisor", "A/S First Name", "A/S Last Name", "A/S EMail",
        "A/S Phone #", "Management Company", "Mgmt Address1", "Mgmt Address2",
        "Mgmt City", "Mgmt State", "Mgmt ZIP", "Mgmt Phone #", "Mgmt Fax #",
        "Corp Mgmt Id", "Property Type", "Property Tags", "Asset or Fee",
        "Owner Name", "Owner Address", "Owner Phone #", "Owner Fax #",
        "PM Software", "PM Software Notes", "Latitude", "Longitude", "API Id",
    ]
    wb = openpyxl.Workbook()
    cover = wb.active
    cover.title = "Cover"
    cover["A1"] = "Some vendor banner — should be ignored by the loader."

    ws = wb.create_sheet("ALN Property Data")
    ws.append(headers)
    for row_dict in rows:
        ws.append([row_dict.get(h) for h in headers])

    path = tmp_path / "fixture-aln.xlsx"
    wb.save(path)
    return path


def _good_row(**overrides) -> dict:
    """A reasonable default ALN row; override fields per test."""
    base = {
        "ALN Id":              "100001",
        "Status":              92,
        "Status Description":  "92-Published",
        "Property Name":       "Test Apartments",
        "Address":             "123 Test St",
        "City":                "Norfolk",
        "State":               "VA",
        "ZIP":                 "23502",
        "County":              "Norfolk",
        "Market":              "NOR",
        "Market Description":  "VA - Norfolk",
        "Submarket":           "NOR-CENTRAL",
        "Lease Terms":         "12",
        "# of Units":          120,
        "Year Built":          "1985",
        "Remodeled":           "2010",
        "Occupancy":           "95",
        "Average Sqft":        850.5,
        "Average Rate":        1500.0,
        "Average Rent/Sqft":   1.7637,
        "ALN Price Class":     "C",
        "Property Type":       "Garden (2)",
        "Property Tags":       "",
        "Asset or Fee":        "Asset",
        "Owner Name":          "Test Owner LLC",
        "Owner Address":       "PO Box 123",
        "Owner Phone #":       "(757) 555-1234",
        "Current Manager":     "Jane Doe",
        "Management Company":  "Test Mgmt Co",
        "PM Software":         "Yardi",
        "Latitude":            "36.9152",
        "Longitude":           "-76.2294",
        "API Id":              "uuid-test-001",
    }
    base.update(overrides)
    return base


# ---------------------------------------------------------------------------
# load_aln_xlsx
# ---------------------------------------------------------------------------


def test_load_xlsx_basic(tmp_path):
    path = _build_fixture_xlsx(tmp_path, [_good_row()])
    rows = load_aln_xlsx(path)
    assert len(rows) == 1
    r = rows[0]
    assert r["property_id"] == "uuid-test-001"
    assert r["aln_id"] == "100001"
    assert r["name"] == "Test Apartments"
    assert r["city"] == "Norfolk"
    assert r["units"] == 120
    assert r["year_built"] == 1985
    assert r["last_remodel"] == 2010
    assert r["occupancy_pct"] == pytest.approx(0.95)
    assert r["avg_sqft"] == pytest.approx(850.5)
    assert r["avg_rent"] == pytest.approx(1500.0)
    assert r["asset_class"] == "C"
    assert r["latitude"] == pytest.approx(36.9152)
    assert r["longitude"] == pytest.approx(-76.2294)
    assert r["asset_or_fee"] == "Asset"
    assert r["aln_pull_date"]  # ISO date string set
    # raw_row preserves the original ALN headers
    raw = json.loads(r["raw_row"])
    assert raw["Property Name"] == "Test Apartments"
    assert raw["API Id"] == "uuid-test-001"


def test_load_skips_rows_missing_required_fields(tmp_path):
    path = _build_fixture_xlsx(tmp_path, [
        _good_row(),
        _good_row(**{"Property Name": None, "API Id": "uuid-no-name"}),
        _good_row(**{"API Id": None}),
    ])
    rows = load_aln_xlsx(path)
    assert len(rows) == 1  # only the good row survives
    assert rows[0]["property_id"] == "uuid-test-001"


def test_load_dedupes_by_property_id(tmp_path):
    """ALN sometimes ships exact-duplicate rows; we keep first, drop rest."""
    path = _build_fixture_xlsx(tmp_path, [
        _good_row(),
        _good_row(),  # same API Id
        _good_row(**{"API Id": "uuid-test-002", "Property Name": "Different"}),
    ])
    rows = load_aln_xlsx(path)
    assert len(rows) == 2
    ids = {r["property_id"] for r in rows}
    assert ids == {"uuid-test-001", "uuid-test-002"}


def test_load_handles_na_occupancy(tmp_path):
    path = _build_fixture_xlsx(tmp_path, [
        _good_row(**{"Occupancy": "N/A"}),
    ])
    rows = load_aln_xlsx(path)
    assert rows[0]["occupancy_pct"] is None


def test_load_handles_empty_rows(tmp_path):
    """Trailing all-empty rows from openpyxl shouldn't blow up the loader."""
    path = _build_fixture_xlsx(tmp_path, [
        _good_row(),
        {h: None for h in [
            "Property Name", "API Id", "City", "Address",
        ]},  # totally empty row
    ])
    rows = load_aln_xlsx(path)
    assert len(rows) == 1


def test_load_missing_file_raises(tmp_path):
    with pytest.raises(FileNotFoundError):
        load_aln_xlsx(tmp_path / "nope.xlsx")


def test_load_missing_sheet_raises(tmp_path):
    """xlsx with the wrong sheet name should raise a clear error."""
    wb = openpyxl.Workbook()
    wb.active.title = "Wrong Name"
    path = tmp_path / "bad.xlsx"
    wb.save(path)
    with pytest.raises(ValueError, match="ALN Property Data"):
        load_aln_xlsx(path)


# ---------------------------------------------------------------------------
# write_to_sqlite
# ---------------------------------------------------------------------------


def _project_schema() -> Path:
    return Path(__file__).resolve().parent.parent / "data" / "schema.sql"


def test_write_creates_table_and_indexes(tmp_path):
    db = tmp_path / "test.db"
    rows = load_aln_xlsx(_build_fixture_xlsx(tmp_path, [_good_row()]))
    n = write_to_sqlite(rows, db, _project_schema())
    assert n == 1

    conn = sqlite3.connect(db)
    try:
        # Table exists with right column count
        cols = conn.execute("PRAGMA table_info(properties)").fetchall()
        col_names = [c[1] for c in cols]
        for required in SCHEMA_COLUMNS:
            assert required in col_names, f"missing column {required!r}"

        # Indexes present
        indexes = {
            r[0] for r in conn.execute(
                "SELECT name FROM sqlite_master WHERE type='index' AND tbl_name='properties'"
            )
        }
        assert "ix_properties_city" in indexes
        assert "ix_properties_class" in indexes
        assert "ix_properties_class_city" in indexes
        assert "ix_properties_latlng" in indexes
    finally:
        conn.close()


def test_write_idempotent_on_repeated_calls(tmp_path):
    """Sync should drop+recreate the table — calling twice should yield same row count."""
    db = tmp_path / "test.db"
    rows = load_aln_xlsx(_build_fixture_xlsx(tmp_path, [_good_row()]))
    write_to_sqlite(rows, db, _project_schema())
    write_to_sqlite(rows, db, _project_schema())  # second run shouldn't error
    conn = sqlite3.connect(db)
    try:
        n = conn.execute("SELECT COUNT(*) FROM properties").fetchone()[0]
        assert n == 1
    finally:
        conn.close()


def test_write_preserves_types(tmp_path):
    """Confirm ints, floats, strings, nulls round-trip correctly through SQLite."""
    db = tmp_path / "test.db"
    rows = load_aln_xlsx(_build_fixture_xlsx(tmp_path, [
        _good_row(**{"Occupancy": "N/A"}),  # null float
    ]))
    write_to_sqlite(rows, db, _project_schema())
    conn = sqlite3.connect(db)
    try:
        row = conn.execute(
            "SELECT units, occupancy_pct, latitude FROM properties"
        ).fetchone()
        assert isinstance(row[0], int)
        assert row[1] is None
        assert isinstance(row[2], float)
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# End-to-end sync
# ---------------------------------------------------------------------------


def test_sync_end_to_end(tmp_path):
    aln = _build_fixture_xlsx(tmp_path, [
        _good_row(),
        _good_row(**{"API Id": "uuid-002", "Property Name": "Second"}),
        _good_row(),  # duplicate of first → dropped
    ])
    db = tmp_path / "test.db"
    # Pass tmp_path as properties_root so the test doesn't pick up the live
    # _custom_props.json from the real Properties/ folder.
    n = sync(aln, db, _project_schema(), properties_root=tmp_path)
    assert n == 2

    conn = sqlite3.connect(db)
    try:
        names = sorted(r[0] for r in conn.execute("SELECT name FROM properties"))
        assert names == ["Second", "Test Apartments"]
    finally:
        conn.close()
