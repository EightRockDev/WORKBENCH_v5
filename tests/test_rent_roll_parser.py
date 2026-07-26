"""Deterministic XLSX/CSV ingestion - no API key required (spec 11).

The bug this locks down: uploading an Excel rent roll failed with
"ANTHROPIC_API_KEY not set" even though a spreadsheet is structured data that
needs no model at all.
"""

from __future__ import annotations

import datetime as dt

import openpyxl
import pytest

from core import document_ingest as di
from core import rent_roll_parser as rrp


def _write_rent_roll_xlsx(path, n_units=26, title_rows=True, total_row=True,
                          headers=("Unit", "Unit Type", "Status", "Resident",
                                   "Sq Ft", "Market Rent", "Actual Rent",
                                   "Lease Expiration")):
    """A realistic property-management export: title block above the table,
    currency strings, a totals row after the units."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rent Roll"
    if title_rows:
        ws.append(["Crossroads Townhomes"])
        ws.append(["Rent Roll as of 06/22/2026"])
        ws.append([])
    ws.append(list(headers))
    for i in range(n_units):
        vacant = i % 13 == 12
        ws.append([
            f"{100 + i}",
            "2BR/1.5BA" if i % 2 else "1BR/1BA",
            "Vacant" if vacant else "Occupied",
            "" if vacant else f"Resident {i}",
            975 if i % 2 else 750,
            "$1,250.00" if i % 2 else "$1,050.00",
            "" if vacant else ("$1,225.00" if i % 2 else "$995.00"),
            "" if vacant else (dt.date(2026, 1 + (i % 12), 15)),
        ])
    if total_row:
        ws.append(["Total", "", "", "", "", "$29,900.00", "$27,000.00", ""])
    wb.save(path)
    return path


def _write_t12_xlsx(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "T-12"
    ws.append(["Crossroads Townhomes - Trailing Twelve Months"])
    ws.append([])
    months = [f"2025-{m:02d}" for m in range(7, 13)] + [f"2026-{m:02d}" for m in range(1, 7)]
    ws.append(["Account"] + months + ["Total"])
    def row(label, monthly, total):
        ws.append([label] + [monthly] * 12 + [total])
    row("Gross Potential Rent", 31_250, 375_000)
    row("Vacancy Loss", -2_200, -26_400)
    row("Other Income", 1_450, 17_400)
    row("Total Revenue", 30_500, 366_000)
    ws.append([])
    row("Payroll", 4_500, 54_000)
    row("Repairs & Maintenance", 2_800, 33_600)
    row("Utilities", 2_100, 25_200)
    row("Management Fee", 1_500, 18_000)
    row("Real Estate Taxes", 2_400, 28_800)
    row("Insurance", 1_100, 13_200)
    row("Total Operating Expenses", 14_400, 172_800)
    row("Net Operating Income", 16_100, 193_200)
    wb.save(path)
    return path


# ------------------------------------------------------------- rent roll

def test_parses_a_realistic_pm_export(tmp_path):
    f = _write_rent_roll_xlsx(tmp_path / "Crossroads Townhomes - Rent Roll - 6.22.26.xlsx")
    block = rrp.parse_rent_roll(f)
    assert block is not None
    assert block["summary"]["totalUnits"] == 26          # total row excluded
    assert block["summary"]["vacant"] == 2
    assert block["summary"]["occupied"] == 24
    u0 = block["units"][0]
    assert u0["unit"] == "100" and u0["status"] == "Occupied"
    assert u0["actualRent"] == 995.0                      # "$995.00" parsed
    assert u0["leaseExp"] == "2026-01-15"


def test_title_rows_and_total_row_are_skipped(tmp_path):
    f = _write_rent_roll_xlsx(tmp_path / "rr.xlsx")
    block = rrp.parse_rent_roll(f)
    labels = [u["unit"] for u in block["units"]]
    assert "Total" not in labels and "Crossroads Townhomes" not in labels


def test_alias_headers_are_recognized(tmp_path):
    f = _write_rent_roll_xlsx(
        tmp_path / "rr2.xlsx",
        headers=("Apt No", "Floorplan", "Lease Status", "Tenant Name",
                 "SF", "Scheduled Rent", "Current Rent", "Lease End"))
    block = rrp.parse_rent_roll(f)
    assert block is not None and block["summary"]["totalUnits"] == 26
    assert block["units"][0]["sqft"] == 750


def test_status_inferred_from_tenant_when_no_status_column(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Unit", "Resident", "Market Rent", "Actual Rent"])
    ws.append(["101", "Jane Doe", 1000, 980])
    ws.append(["102", "", 1000, ""])
    f = tmp_path / "rr3.xlsx"
    wb.save(f)
    block = rrp.parse_rent_roll(f)
    statuses = {u["unit"]: u["status"] for u in block["units"]}
    assert statuses == {"101": "Occupied", "102": "Vacant"}


def test_csv_works_too(tmp_path):
    f = tmp_path / "rentroll.csv"
    f.write_text("Unit,Status,Market Rent,Actual Rent\n"
                 "A1,Occupied,\"$1,100\",\"$1,075\"\n"
                 "A2,Vacant,\"$1,100\",\n")
    block = rrp.parse_rent_roll(f)
    assert block["summary"]["totalUnits"] == 2
    assert block["units"][0]["actualRent"] == 1075.0


def test_unrecognizable_sheet_returns_none_not_garbage(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Quarter", "Revenue", "Margin"])
    ws.append(["Q1", 100, 0.2])
    f = tmp_path / "not_a_rent_roll.xlsx"
    wb.save(f)
    assert rrp.parse_rent_roll(f) is None


# ------------------------------------------------------------------ T-12

def test_parses_a_t12_taking_the_totals_column(tmp_path):
    f = _write_t12_xlsx(tmp_path / "T12.xlsx")
    out = rrp.parse_t12(f)
    assert out is not None
    assert out["totalRevenue"] == 366_000
    assert out["totalOpex"] == 172_800
    assert out["noi"] == 193_200
    assert out["t12_revenue"]["grossPotentialRent"] == 375_000
    assert out["t12_revenue"]["vacancy"] == 26_400        # sign normalized
    assert out["t12_fixedCharges"]["insurance"] == 13_200


def test_t12_backbone_required(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Payroll", 54_000])       # expense lines but no revenue
    ws.append(["Utilities", 25_200])
    f = tmp_path / "fragment.xlsx"
    wb.save(f)
    assert rrp.parse_t12(f) is None


# --------------------------------------------------- ingest_document routing

def test_xlsx_rent_roll_ingests_without_api_key(tmp_path, monkeypatch):
    """THE reported bug: Excel rent roll + no ANTHROPIC_API_KEY must succeed."""
    monkeypatch.delenv("ANTHROPIC_API_KEY", raising=False)
    f = _write_rent_roll_xlsx(tmp_path / "Crossroads Townhomes - Rent Roll - 6.22.26.xlsx")
    result = di.ingest_document(f)
    assert result.error is None, result.error
    assert result.is_success
    assert result.document_type == "rent_roll"
    assert result.extracted["rentRoll"]["summary"]["totalUnits"] == 26
    assert "no AI used" in result.extraction_notes


def test_xlsx_t12_ingests_without_api_key(tmp_path, monkeypatch):
    monkeypatch.delenv("ANTHROPIC_API_KEY", raising=False)
    f = _write_t12_xlsx(tmp_path / "Crossroads T-12.xlsx")
    result = di.ingest_document(f)
    assert result.error is None and result.document_type == "t12"
    assert result.extracted["noi"] == 193_200


def test_unrecognized_xlsx_without_key_gives_actionable_error(tmp_path, monkeypatch):
    monkeypatch.delenv("ANTHROPIC_API_KEY", raising=False)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Weird Layout Rent Roll"])   # classified rent_roll by filename,
    ws.append(["no", "usable", "headers"])  # but not parseable
    f = tmp_path / "weird rent roll.xlsx"
    wb.save(f)
    result = di.ingest_document(f)
    assert not result.is_success
    assert result.error.startswith("NEEDS_API_KEY")
    assert "not recognized" in result.error


def test_pdf_without_key_gives_actionable_error(tmp_path, monkeypatch):
    monkeypatch.delenv("ANTHROPIC_API_KEY", raising=False)
    # A tiny valid one-page PDF with enough text to classify as a rent roll.
    import pypdf
    writer = pypdf.PdfWriter()
    writer.add_blank_page(width=612, height=792)
    f = tmp_path / "rent roll scan.pdf"
    with f.open("wb") as fh:
        writer.write(fh)
    result = di.ingest_document(f, document_type="rent_roll")
    assert not result.is_success
    assert result.error.startswith("NEEDS_API_KEY")


def test_commit_roundtrip_feeds_the_rent_roll_ui(tmp_path, monkeypatch):
    """Parsed output must land in sources.json in the exact shape
    ui/rent_roll.py and the anomaly detectors read."""
    monkeypatch.delenv("ANTHROPIC_API_KEY", raising=False)
    f = _write_rent_roll_xlsx(tmp_path / "rent roll.xlsx")
    result = di.ingest_document(f)
    folder = tmp_path / "prop"
    folder.mkdir()
    n = di.commit_to_sources_json(folder, result, overwrite=True)
    assert n > 0

    from data.property_io import load_sources
    from core.rent_roll_anomalies import detect_anomalies
    from core.extraction_qa import run_qa

    sources = load_sources(folder)
    rr = sources["rentRoll"]
    # The rentRoll block must be stored RAW (no provenance wrapping) because
    # ui/rent_roll.py reads summary values and unit fields directly.
    assert rr["summary"]["totalUnits"] == 26          # raw int, not {"value": ...}
    assert isinstance(rr["summary"]["occupied"], int)
    assert len(rr["units"]) == 26
    assert rr["units"][0]["actualRent"] == 995.0
    # and the QA/anomaly layers accept it without blowing up
    detect_anomalies(sources)
    report = run_qa(sources)
    assert not report.errors, [c.detail for c in report.errors]


# ----------------------------------------------- hostile-file resilience

def test_zero_byte_upload_gives_a_plain_explanation(tmp_path, monkeypatch):
    """THE reported bug (round 2): a 0-byte upload (cloud-only OneDrive
    placeholder / drag-from-email) must produce a human explanation, never
    'could not extract text from ...'."""
    monkeypatch.delenv("ANTHROPIC_API_KEY", raising=False)
    f = tmp_path / "Crossroads Townhomes - T12 - 05.2026 - Corrected.xlsx"
    f.write_bytes(b"")
    result = di.ingest_document(f)
    assert not result.is_success
    assert result.error.startswith("EMPTY_FILE")
    assert "OneDrive" in result.error and "0 bytes" in result.error


def test_garbage_bytes_named_xlsx_fail_cleanly(tmp_path, monkeypatch):
    monkeypatch.delenv("ANTHROPIC_API_KEY", raising=False)
    f = tmp_path / "corrupt rent roll.xlsx"
    f.write_bytes(b"this is not a zip archive at all" * 10)
    result = di.ingest_document(f)          # must not raise
    assert not result.is_success
    assert "could not read" in result.error
    assert result.error != "could not extract text from corrupt rent roll.xlsx"


def test_xlsx_bytes_with_xls_extension_still_parse(tmp_path, monkeypatch):
    """PM systems mislabel exports; the reader chain must recover."""
    monkeypatch.delenv("ANTHROPIC_API_KEY", raising=False)
    real = _write_rent_roll_xlsx(tmp_path / "real.xlsx")
    mislabeled = tmp_path / "rent roll export.xls"
    mislabeled.write_bytes(real.read_bytes())
    result = di.ingest_document(mislabeled)
    assert result.is_success, result.error
    assert result.extracted["rentRoll"]["summary"]["totalUnits"] == 26


def test_empty_file_beats_every_other_error_path(tmp_path, monkeypatch):
    """Empty PDF and empty CSV get the same explanation - the check runs
    before any format-specific reader."""
    monkeypatch.delenv("ANTHROPIC_API_KEY", raising=False)
    for name in ("t12.pdf", "rent roll.csv", "om.xlsx"):
        f = tmp_path / name
        f.write_bytes(b"")
        result = di.ingest_document(f)
        assert result.error.startswith("EMPTY_FILE"), name
