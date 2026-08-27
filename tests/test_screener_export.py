"""Property Screener downloads — CSV and Excel of the rows on screen.

The owner's ask (2026-08-27) is a file he can slice, so the tests care
about the two ways an export betrays that: a numeric column that isn't
numeric (the screen's "—" leaking into the file), and a file that
disagrees with the screen it was taken from.
"""

from __future__ import annotations

import csv
import datetime as _dt
import io

from openpyxl import load_workbook

from core.screener import SOURCE_COUNTY, SOURCE_CURATED
from core.screener_export import (HEADERS, build_csv_bytes, build_xlsx_bytes,
                                  export_filename, to_records)

CURATED = {
    "source": SOURCE_CURATED, "property_id": "P1",
    "name": "Grand Hampton at Langley", "address": "611 Michigan Dr",
    "city": "Hampton", "state": "VA", "zip": "23669", "units": 192,
    "year_built": 1967, "asset_class": "C", "market": "Hampton Roads",
    "owner": "Dolly Parton Holdings LLC", "management_company": "Dollywood Mgmt",
    "occupancy_pct": 0.94, "sale_price": 21_500_000, "sale_when": "2019",
}
# A county row: no name, class, management company or occupancy.
COUNTY = {
    "source": SOURCE_COUNTY, "property_id": "8R-B", "name": "1 BROAD ST",
    "address": "1 BROAD ST", "city": "Richmond", "state": "VA",
    "zip": "07030", "units": 40, "year_built": None, "asset_class": None,
    "market": None, "owner": "BROAD ST LLC", "management_company": None,
    "occupancy_pct": None, "sale_price": None, "sale_when": None,
}
ROWS = [CURATED, COUNTY]


def _csv_rows(data: bytes) -> list[list[str]]:
    return list(csv.reader(io.StringIO(data.decode("utf-8-sig"))))


def test_csv_carries_every_row_under_the_shared_header():
    rows = _csv_rows(build_csv_bytes(ROWS))
    assert rows[0] == list(HEADERS)
    assert len(rows) == 3
    assert rows[1][HEADERS.index("Name")] == "Grand Hampton at Langley"
    assert rows[2][HEADERS.index("Owner")] == "BROAD ST LLC"


def test_missing_values_are_blank_not_the_screen_dash():
    """"—" in a numeric column is what stops a spreadsheet being a model."""
    body = build_csv_bytes(ROWS).decode("utf-8-sig")
    assert "—" not in body
    county = _csv_rows(build_csv_bytes([COUNTY]))[1]
    assert county[HEADERS.index("Last Sale Price")] == ""
    assert county[HEADERS.index("Occupancy %")] == ""


def test_money_exports_as_a_number_not_formatted_text():
    row = _csv_rows(build_csv_bytes([CURATED]))[1]
    assert row[HEADERS.index("Last Sale Price")] == "21500000"


def test_occupancy_fraction_becomes_percent():
    assert to_records([CURATED])[0]["Occupancy %"] == 94.0


def test_export_adds_the_columns_the_screen_has_no_room_for():
    rec = to_records([CURATED])[0]
    assert rec["Address"] == "611 Michigan Dr"
    assert rec["Market"] == "Hampton Roads"
    assert rec["Property ID"] == "P1"


def test_csv_opens_cleanly_in_excel():
    assert build_csv_bytes(ROWS).startswith(b"\xef\xbb\xbf")


def _sheet(data: bytes, name: str):
    return load_workbook(io.BytesIO(data))[name]


def test_xlsx_results_sheet_matches_the_rows():
    ws = _sheet(build_xlsx_bytes(ROWS), "Results")
    assert [c.value for c in ws[1]] == list(HEADERS)
    assert ws.max_row == 3
    assert ws.cell(row=2, column=HEADERS.index("Units") + 1).value == 192
    assert ws.cell(row=2,
                   column=HEADERS.index("Last Sale Price") + 1).value == 21_500_000
    assert ws.freeze_panes == "A2"


def test_zip_stays_text_so_leading_zeros_survive():
    ws = _sheet(build_xlsx_bytes([COUNTY]), "Results")
    cell = ws.cell(row=2, column=HEADERS.index("Zip") + 1)
    assert cell.value == "07030"
    assert cell.number_format == "@"


def test_blank_cells_stay_empty_in_the_workbook():
    ws = _sheet(build_xlsx_bytes([COUNTY]), "Results")
    assert ws.cell(row=2, column=HEADERS.index("Class") + 1).value is None
    assert ws.cell(row=2,
                   column=HEADERS.index("Last Sale Price") + 1).value is None


def test_filters_sheet_names_the_search_that_produced_the_file():
    filters = {"city": "Hampton", "units_min": 60, "units_max": None,
               "asset_class": ["C", "D"], "name": "", "occ_min": None}
    ws = _sheet(build_xlsx_bytes(ROWS, filters=filters,
                                generated_at=_dt.datetime(2026, 8, 27, 9, 5)),
                "Filters")
    text = {(r[0].value, r[1].value) for r in ws.iter_rows(min_row=1,
                                                           max_col=2)}
    assert ("City", "Hampton") in text
    assert ("Units — min", "60") in text
    assert ("Class", "C, D") in text
    assert ("Rows", 2) in text
    assert ("Generated", "2026-08-27 09:05") in text
    # An inactive filter must not appear — a blank line reads as a filter
    # that was applied and matched nothing.
    assert not any(lbl in ("Property name", "Occupancy % — min",
                           "Units — max")
                   for lbl, _v in text if lbl)


def test_no_filters_says_so_rather_than_leaving_the_sheet_bare():
    ws = _sheet(build_xlsx_bytes(ROWS, filters={}), "Filters")
    assert any("none" in str(c.value) for c in ws["A"] if c.value)


def test_empty_result_set_still_produces_a_readable_file():
    assert _csv_rows(build_csv_bytes([]))[0] == list(HEADERS)
    ws = _sheet(build_xlsx_bytes([]), "Results")
    assert ws.max_row == 1
    assert ws.auto_filter.ref is None


def test_filename_is_stamped_and_typed():
    name = export_filename("xlsx", when=_dt.datetime(2026, 8, 27, 9, 5))
    assert name == "property-screener-20260827-0905.xlsx"


def test_the_screen_offers_both_downloads_from_the_rows_it_painted():
    """The export must read the rendered list, never re-run the query."""
    src = open("ui/property_screener.py", encoding="utf-8").read()
    assert "build_csv_bytes(rows)" in src
    assert "build_xlsx_bytes(rows" in src
    # _render_downloads takes the same `rows` _render_results paints.
    assert "_render_downloads(c, rows)" in src
    assert src.count("run_screener(") == 1  # Submit, and nowhere else


# --- The screen itself, rendered ------------------------------------------
# The grep above proves the wiring; these run the real Streamlit widgets,
# because "the deliverable is a screen" (CLAUDE.md, 2026-08-15).

import textwrap  # noqa: E402

from streamlit.testing.v1 import AppTest  # noqa: E402

_SCRIPT = textwrap.dedent("""
    import sys
    sys.path.insert(0, %r)
    import config
    from ui.property_screener import _render_results
    from tests.test_screener_export import CURATED, COUNTY
    import streamlit as st
    n = st.session_state.get("_n", 2)
    rows = ([CURATED, COUNTY] * (n // 2))[:n]
    _render_results(config.COLORS, rows)
""") % str(__import__("pathlib").Path(__file__).resolve().parent.parent)


def _run(n=2):
    at = AppTest.from_string(_SCRIPT, default_timeout=60)
    at.session_state["_n"] = n
    return at.run()


def test_the_results_screen_renders_both_download_buttons():
    at = _run()
    assert not at.exception
    buttons = at.get("download_button")
    assert [b.proto.label for b in buttons] == ["Download CSV", "Download Excel"]
    assert buttons[0].proto.url.endswith(".csv")
    assert buttons[1].proto.url.endswith(".xlsx")


def test_no_results_means_no_download_buttons():
    at = _run(0)
    assert not at.get("download_button")
    assert at.info


def test_a_capped_result_set_says_the_file_is_capped_too():
    from core.screener import DEFAULT_LIMIT

    body = " ".join(str(m.value) for m in _run(DEFAULT_LIMIT).markdown)
    assert "caps results at" in body
    assert "caps results at" not in " ".join(
        str(m.value) for m in _run(2).markdown)
