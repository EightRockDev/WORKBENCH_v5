"""Tests for core.excel_export — the deterministic 3-sheet deal workbook.

Builds a `data` dict the same way `ui.exec_summary._build_summary_data` does
(real cash-flow projection + verdict), then round-trips the workbook bytes
through openpyxl to confirm the sheets, headers, and key values are present.
No folder/storage I/O — the export takes plain dicts.
"""

from io import BytesIO

from openpyxl import load_workbook

import config
from core.calc import DebtTerms, build_cashflow, build_debt_schedule
from core.excel_export import build_workbook_bytes
from core.irr import equity_multiple, lp_irr, project_irr
from core.verdict import evaluate
from core.waterfall import run_waterfall
from data.property_io import DealState


def _mk_deal(**kw) -> DealState:
    base = dict(pp=5_000_000, noi=350_000, dp=30, ir=6.5, vac=8, rg=3,
                eg=3, xc=6.5, hp=5)
    base.update(kw)
    return DealState(**base)


def _mk_data(deal: DealState, prop: dict, sources: dict | None = None) -> dict:
    er = config.EXPENSE_RATIOS.get("C", 0.45)
    denom = 1.0 - deal.vacancy_frac - er
    gpr = deal.noi / denom if denom > 0 else deal.noi / 0.5
    expenses = gpr * er

    debt_terms = DebtTerms(
        loan_amount=deal.loan_amount, annual_rate=deal.interest_rate,
        amort_months=config.AMORT_MONTHS, io_years=deal.io,
    )
    debt_sched = build_debt_schedule(debt_terms, deal.hp)
    cf = build_cashflow(
        year1_gpr=gpr, year1_vacancy_pct=deal.vacancy_frac,
        year1_expenses=expenses, rent_growth=deal.rent_growth,
        expense_growth=deal.expense_growth, am_fee_pct=deal.am_fee_pct,
        debt=debt_sched, hold_years=deal.hp, exit_cap=deal.exit_cap,
        equity_raise=deal.equity_raise,
    )
    pots = [r.cash_flow for r in cf.rows[:-1]] + [
        cf.rows[-1].cash_flow + cf.exit_proceeds_net]
    wf = run_waterfall(equity_raise=deal.equity_raise, annual_pots=pots)
    cap = deal.noi / deal.pp if deal.pp else 0.0
    ads = debt_sched.annual_payment[0]
    dscr_v = (deal.noi - deal.am_fee_pct * gpr) / ads if ads else 0.0
    ppu = deal.pp / (prop.get("units") or 1)
    verdict = evaluate(cap=cap, dscr=dscr_v, coc=cf.rows[0].coc, ppu=ppu,
                       city=prop.get("city") or "")
    return {
        "deal": deal, "sources": sources, "cf": cf, "wf": wf,
        "lp_irr": lp_irr(wf.lp_cashflows),
        "project_irr": project_irr(
            equity_raise=deal.equity_raise,
            annual_cashflows=[r.cash_flow for r in cf.rows],
            exit_proceeds_net=cf.exit_proceeds_net),
        "equity_multiple": equity_multiple(deal.equity_raise,
                                           wf.total_lp_distributions),
        "cap": cap, "dscr": dscr_v, "coc": cf.rows[0].coc, "ppu": ppu,
        "verdict": verdict, "tightened": None,
    }


_PROP = {"name": "Madison Terrace", "address": "1 Main St",
         "city": "Norfolk", "state": "VA", "units": 100, "year_built": 1978}


def _open(xlsx: bytes):
    return load_workbook(BytesIO(xlsx))


def test_workbook_has_three_named_sheets():
    data = _mk_data(_mk_deal(), _PROP)
    wb = _open(build_workbook_bytes(_PROP, data, None))
    assert wb.sheetnames == ["Summary", "Returns", "Rent Roll"]


def test_summary_sheet_carries_property_and_price():
    data = _mk_data(_mk_deal(), _PROP)
    wb = _open(build_workbook_bytes(_PROP, data, None))
    ws = wb["Summary"]
    cells = [c.value for row in ws.iter_rows() for c in row]
    assert "Madison Terrace — Executive Summary" in cells
    assert 5_000_000 in cells               # asking price
    assert any(v == data["verdict"].verdict for v in cells)


def test_returns_sheet_has_a_row_per_hold_year_plus_exit():
    deal = _mk_deal(hp=5)
    data = _mk_data(deal, _PROP)
    wb = _open(build_workbook_bytes(_PROP, data, None))
    ws = wb["Returns"]
    col_a = [c.value for c in ws["A"]]
    for yr in range(1, 6):
        assert yr in col_a
    assert "Exit & returns" in col_a


def test_rent_roll_note_when_no_rent_roll():
    data = _mk_data(_mk_deal(), _PROP)
    wb = _open(build_workbook_bytes(_PROP, data, None))
    ws = wb["Rent Roll"]
    texts = [c.value for row in ws.iter_rows() for c in row if c.value]
    assert any("No rent roll ingested" in str(t) for t in texts)


def test_rent_roll_units_are_written():
    sources = {"rentRoll": {
        "file": "rr.xlsx", "date": "2026-07-01",
        "summary": {"totalUnits": 2, "occupiedUnits": 1,
                    "totalMarketRent": 2500.0, "totalActualRent": 1200.0},
        "units": [
            {"unit": "101", "unitType": "1BR", "status": "Occupied",
             "tenant": "Smith", "sqft": 700, "marketRent": 1300.0,
             "actualRent": 1200.0, "moveIn": "2024-01-01",
             "leaseExp": "2026-12-31"},
            {"unit": "102", "unitType": "1BR", "status": "Vacant",
             "tenant": None, "sqft": 700, "marketRent": 1200.0,
             "actualRent": None, "moveIn": None, "leaseExp": None},
        ],
    }}
    data = _mk_data(_mk_deal(), _PROP, sources)
    wb = _open(build_workbook_bytes(_PROP, data, sources))
    ws = wb["Rent Roll"]
    col_a = [c.value for c in ws["A"]]
    assert "101" in col_a
    assert "102" in col_a
    assert "Total units" in col_a


def test_bytes_are_a_valid_xlsx_zip():
    data = _mk_data(_mk_deal(), _PROP)
    b = build_workbook_bytes(_PROP, data, None)
    assert b[:2] == b"PK"          # xlsx is a zip archive
    assert len(b) > 2000
